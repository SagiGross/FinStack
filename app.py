"""
FinStack CFO v6 — Auth + CRM + Sales Pipeline
================================================
INSTALL:  py -m pip install fastapi uvicorn pyjwt
RUN:      py app.py
"""
from __future__ import annotations
import sqlite3, json, os, time, hashlib, secrets
from contextlib import contextmanager
from datetime import date, timedelta, datetime
from pathlib import Path
from typing import Optional
from uuid import uuid4

import jwt
from fastapi import FastAPI, Query, Header, HTTPException, Depends, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# Simple password hashing (no bcrypt dependency needed)
def hash_pw(password: str) -> str:
    salt = secrets.token_hex(16)
    h = hashlib.sha256((salt + password).encode()).hexdigest()
    return salt + ":" + h

def verify_pw(password: str, stored: str) -> bool:
    if ":" not in stored: return False
    salt, h = stored.split(":", 1)
    return hashlib.sha256((salt + password).encode()).hexdigest() == h

app = FastAPI(title="FinStack CFO API", version="6.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

ECF = 1.35; FY = 2026  # ECF kept as fallback

# Israeli Social Security brackets (2026) - EMPLOYER ONLY
SS_LOW_THRESHOLD = 7703    # ILS
SS_HIGH_THRESHOLD = 51910  # ILS
SS_LOW_RATE = 0.0451       # employer: 4.51% on low bracket (leumi only, no briut)
SS_HIGH_RATE = 0.076       # employer: 7.6% on high bracket

def calc_social_security(gross_ils):
    """Calculate employer social security (bituach leumi + briut) in ILS."""
    if gross_ils <= 0: return 0
    low_part = min(gross_ils, SS_LOW_THRESHOLD)
    high_part = max(0, min(gross_ils, SS_HIGH_THRESHOLD) - SS_LOW_THRESHOLD)
    return low_part * SS_LOW_RATE + high_part * SS_HIGH_RATE

def calc_employer_cost(gross, bonus=0, position_pct=100, pension_pct=8.33, disability_pct=2.5, study_fund_pct=7.5, study_fund_salary=0, currency='ILS'):
    """Calculate total monthly employer cost in original currency."""
    base = gross * (position_pct / 100.0)
    # Social security (only for ILS employees)
    if currency == 'ILS':
        ss = calc_social_security(base)
    else:
        ss = base * 0.15  # rough estimate for non-ILS
    # Pension & disability on gross
    pension = base * (pension_pct / 100.0)
    disability = base * (disability_pct / 100.0)
    # Study fund on insured salary (or gross if not set)
    sf_base = study_fund_salary if study_fund_salary > 0 else base
    study_fund = sf_base * (study_fund_pct / 100.0)
    # Total
    total = base + ss + pension + disability + study_fund + bonus
    return total, {"gross": base, "socialSecurity": round(ss, 2), "pension": round(pension, 2), 
                   "disability": round(disability, 2), "studyFund": round(study_fund, 2), "bonus": bonus}
ML = [f"{m:02d}-{FY}" for m in range(1, 13)]
JWT_SECRET = os.environ.get("JWT_SECRET", "finstack-secret-key-change-in-production")
STAGES = ["Initial Contact", "Email Sent", "Demo", "Negotiation", "Signed"]
DB_PATH = Path(__file__).parent / "finstack.db"

# ══════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════
@contextmanager
def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try: yield conn; conn.commit()
    finally: conn.close()

def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY, username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL, role TEXT NOT NULL DEFAULT 'sales',
            display_name TEXT, dept TEXT DEFAULT 'S&M',
            can_see_all INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY, value TEXT
        );
        CREATE TABLE IF NOT EXISTS monthly_rates (
            month TEXT PRIMARY KEY, rate REAL NOT NULL, locked INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS employees (
            id TEXT PRIMARY KEY, name TEXT NOT NULL, dept TEXT NOT NULL,
            gross REAL NOT NULL, position_pct REAL DEFAULT 100,
            bonus REAL DEFAULT 0, currency TEXT DEFAULT 'ILS',
            pension_pct REAL DEFAULT 8.33,
            disability_pct REAL DEFAULT 2.5,
            study_fund_pct REAL DEFAULT 7.5,
            study_fund_salary REAL DEFAULT 0,
            hire_date TEXT NOT NULL, term_date TEXT, is_ghost INTEGER DEFAULT 0,
            pay_net_date TEXT DEFAULT '',
            pay_tax_date TEXT DEFAULT '',
            pay_pension_date TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS contracts (
            id TEXT PRIMARY KEY, client TEXT NOT NULL, country TEXT DEFAULT '',
            industry TEXT DEFAULT 'Enterprise', yr INTEGER, sm INTEGER, sd TEXT,
            val REAL NOT NULL, dso INTEGER DEFAULT 60, pm TEXT,
            monthly INTEGER DEFAULT 0, chance REAL, ap REAL,
            currency TEXT DEFAULT 'USD',
            stage TEXT DEFAULT 'Initial Contact', stage_updated_at TEXT,
            salesperson_id TEXT, notes TEXT DEFAULT '[]',
            contact_name TEXT DEFAULT '', contact_phone TEXT DEFAULT '',
            contact_email TEXT DEFAULT '', contact_linkedin TEXT DEFAULT '',
            payment_splits TEXT DEFAULT '[]',
            payment_method TEXT DEFAULT '',
            is_new_client INTEGER DEFAULT 1,
            invoice_date TEXT DEFAULT '',
            subject_to_vat INTEGER DEFAULT 0,
            created_at TEXT, updated_at TEXT,
            FOREIGN KEY (salesperson_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS expenses (
            id TEXT PRIMARY KEY, dept TEXT NOT NULL, vendor TEXT NOT NULL,
            sub_cat TEXT DEFAULT '', is_cogs INTEGER DEFAULT 0, amounts TEXT DEFAULT '{}',
            currency TEXT DEFAULT 'ILS',
            vendor_contact TEXT DEFAULT '', vendor_email TEXT DEFAULT '',
            vendor_phone TEXT DEFAULT '', service_desc TEXT DEFAULT '',
            vendor_bank_num TEXT DEFAULT '', vendor_branch_num TEXT DEFAULT '',
            vendor_account_num TEXT DEFAULT '',
            is_fixed INTEGER DEFAULT 1,
            frequency TEXT DEFAULT 'monthly',
            payment_method TEXT DEFAULT '',
            payment_dates TEXT DEFAULT '{}',
            expense_invoice_date TEXT DEFAULT '',
            expense_payment_date TEXT DEFAULT '',
            is_recurring INTEGER DEFAULT 1,
            subject_to_vat INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS invoices (
            id TEXT PRIMARY KEY,
            expense_id TEXT NOT NULL,
            invoice_number TEXT DEFAULT '',
            amount REAL NOT NULL DEFAULT 0,
            currency TEXT DEFAULT 'ILS',
            due_date TEXT,
            paid_date TEXT,
            paid_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'pending',
            notes TEXT DEFAULT '',
            created_at TEXT,
            FOREIGN KEY (expense_id) REFERENCES expenses(id)
        );
        CREATE TABLE IF NOT EXISTS bank_balances (
            month TEXT,
            bank TEXT DEFAULT 'total',
            opening_balance REAL,
            closing_balance REAL,
            is_manual INTEGER DEFAULT 0,
            notes TEXT DEFAULT '',
            PRIMARY KEY (month, bank)
        );
        CREATE TABLE IF NOT EXISTS payment_methods (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            bank_group TEXT DEFAULT '',
            is_active INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS loans (
            id TEXT PRIMARY KEY,
            lender TEXT NOT NULL,
            amount REAL NOT NULL,
            term_months INTEGER NOT NULL,
            annual_rate REAL DEFAULT 0,
            start_date TEXT NOT NULL,
            repayment_type TEXT DEFAULT 'principal_interest',
            currency TEXT DEFAULT 'ILS',
            notes TEXT DEFAULT '',
            interest_start_month TEXT DEFAULT '',
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS cap_table (
            id TEXT PRIMARY KEY,
            holder_name TEXT NOT NULL,
            shares REAL NOT NULL DEFAULT 0,
            is_esop INTEGER DEFAULT 0,
            notes TEXT DEFAULT '',
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS income_sources (
            id TEXT PRIMARY KEY,
            client TEXT NOT NULL,
            category TEXT DEFAULT '',
            dept TEXT DEFAULT 'Revenue',
            currency TEXT DEFAULT 'ILS',
            contact_name TEXT DEFAULT '',
            contact_phone TEXT DEFAULT '',
            contact_email TEXT DEFAULT '',
            is_recurring INTEGER DEFAULT 0,
            subject_to_vat INTEGER DEFAULT 0,
            notes TEXT DEFAULT '',
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS income_payments (
            id TEXT PRIMARY KEY,
            income_id TEXT NOT NULL,
            amount REAL DEFAULT 0,
            invoice_date TEXT DEFAULT '',
            payment_date TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS user_preferences (
            user_id TEXT NOT NULL,
            pref_key TEXT NOT NULL,
            pref_value TEXT NOT NULL,
            PRIMARY KEY (user_id, pref_key)
        );
        CREATE TABLE IF NOT EXISTS action_items (
            id TEXT PRIMARY KEY,
            contract_id TEXT NOT NULL,
            user_id TEXT NOT NULL,
            description TEXT NOT NULL,
            due_date TEXT NOT NULL,
            completed INTEGER DEFAULT 0,
            completed_at TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS salary_snapshots (
            id TEXT PRIMARY KEY,
            employee_id TEXT NOT NULL,
            month TEXT NOT NULL,
            gross REAL NOT NULL,
            position_pct REAL DEFAULT 100,
            bonus REAL DEFAULT 0,
            currency TEXT DEFAULT 'ILS',
            dept TEXT DEFAULT 'R&D',
            pension_pct REAL DEFAULT 8.33,
            disability_pct REAL DEFAULT 2.5,
            study_fund_pct REAL DEFAULT 7.5,
            study_fund_salary REAL DEFAULT 0,
            net_pay_date TEXT DEFAULT '',
            tax_pay_date TEXT DEFAULT '',
            pension_pay_date TEXT DEFAULT '',
            created_at TEXT,
            UNIQUE(employee_id, month)
        );
        CREATE TABLE IF NOT EXISTS expense_payments (
            id TEXT PRIMARY KEY,
            expense_id TEXT NOT NULL,
            month TEXT NOT NULL,
            amount REAL DEFAULT 0,
            invoice_date TEXT DEFAULT '',
            payment_date TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS alert_dismissals (
            id TEXT PRIMARY KEY,
            alert_key TEXT NOT NULL,
            dismissed_by TEXT NOT NULL,
            dismissed_at TEXT,
            action_note TEXT DEFAULT '',
            UNIQUE(alert_key, dismissed_by)
        );
        CREATE TABLE IF NOT EXISTS sales_targets (
            id TEXT PRIMARY KEY,
            salesperson_id TEXT NOT NULL,
            year INTEGER DEFAULT 2026,
            month INTEGER,
            revenue_target REAL DEFAULT 0,
            deals_target INTEGER DEFAULT 0,
            avg_deal_target REAL DEFAULT 0,
            cycle_days_target INTEGER DEFAULT 0,
            close_rate_target REAL DEFAULT 0,
            new_clients_target INTEGER DEFAULT 0,
            created_at TEXT,
            FOREIGN KEY (salesperson_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS scenarios (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            created_at TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS scenario_items (
            id TEXT PRIMARY KEY,
            scenario_id TEXT NOT NULL,
            item_type TEXT NOT NULL,
            config TEXT DEFAULT '{}',
            FOREIGN KEY (scenario_id) REFERENCES scenarios(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, entity_type TEXT, entity_id TEXT,
            field TEXT, old_value TEXT, new_value TEXT, changed_by TEXT, changed_at TEXT
        );
        """)

def carry_forward(known):
    result = {}; last = 0.0
    for m in ML:
        if m in known and known[m] is not None: last = known[m]
        result[m] = last
    return result

def get_vat_rate():
    """Get VAT rate from settings. Default 18% for Israel 2026."""
    try:
        with get_db() as db:
            r = db.execute("SELECT value FROM app_settings WHERE key='vat_rate'").fetchone()
            if r: return float(r["value"]) / 100
    except: pass
    return 0.18

def get_exchange_rate():
    """Get current USD/ILS rate. Try live API, fallback to stored."""
    # Try stored rate first
    with get_db() as db:
        r = db.execute("SELECT value FROM app_settings WHERE key='usd_ils_rate'").fetchone()
        stored = float(r["value"]) if r else 3.12
    # Try to fetch live rate (cache for 1 hour)
    with get_db() as db:
        cache = db.execute("SELECT value FROM app_settings WHERE key='rate_cache_time'").fetchone()
        cache_time = float(cache["value"]) if cache else 0
    import time as _time
    if _time.time() - cache_time < 3600:  # Use cache if less than 1 hour old
        return stored
    try:
        import urllib.request
        req = urllib.request.Request("https://api.frankfurter.app/latest?from=USD&to=ILS",
                                     headers={"User-Agent":"FinStack/1.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
            rate = data["rates"]["ILS"]
            with get_db() as db:
                db.execute("INSERT OR REPLACE INTO app_settings VALUES ('usd_ils_rate',?)", (str(rate),))
                db.execute("INSERT OR REPLACE INTO app_settings VALUES ('rate_cache_time',?)", (str(_time.time()),))
            return rate
    except:
        return stored

def set_exchange_rate(rate):
    import time as _time
    with get_db() as db:
        db.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('usd_ils_rate',?)", (str(rate),))
        db.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('rate_cache_time',?)", (str(_time.time()),))

def get_rate_for_month(month_str):
    """Locked months use stored rate. Current/future use live."""
    with get_db() as db:
        r = db.execute("SELECT rate, locked FROM monthly_rates WHERE month=?", (month_str,)).fetchone()
        if r and r["locked"]:
            return r["rate"]
    return get_exchange_rate()

def lock_past_months():
    """Lock rates for past months on startup. Uses stored rate if available, otherwise current live rate."""
    now = datetime.now()
    current_month = f"{now.month:02d}-{now.year}"
    live_rate = get_exchange_rate()
    with get_db() as db:
        for m in ML:
            if m < current_month:
                existing = db.execute("SELECT rate, locked FROM monthly_rates WHERE month=?", (m,)).fetchone()
                if not existing:
                    # No rate stored yet — lock with current live rate as best estimate
                    db.execute("INSERT INTO monthly_rates (month, rate, locked) VALUES (?,?,1)", (m, live_rate))
                    print(f"  Rate locked: {m} = {live_rate:.2f} (live fallback)")
                elif not existing["locked"]:
                    # Rate was stored but not locked — lock it (keeps the stored rate)
                    db.execute("UPDATE monthly_rates SET locked=1 WHERE month=?", (m,))
                    print(f"  Rate locked: {m} = {existing['rate']:.2f} (stored)")
                # If already locked, don't touch it

def get_eur_rate():
    """Get EUR/USD rate. Try cached, then live."""
    with get_db() as db:
        r = db.execute("SELECT value FROM app_settings WHERE key='eur_usd_rate'").fetchone()
        stored = float(r["value"]) if r else 1.08
    import time as _t2
    with get_db() as db:
        cache = db.execute("SELECT value FROM app_settings WHERE key='eur_cache_time'").fetchone()
        ct = float(cache["value"]) if cache else 0
    if _t2.time() - ct < 3600: return stored
    try:
        import urllib.request
        req = urllib.request.Request("https://api.frankfurter.app/latest?from=EUR&to=USD",
                                     headers={"User-Agent":"FinStack/1.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
            rate = data["rates"]["USD"]  # 1 EUR = X USD
            with get_db() as db:
                db.execute("INSERT OR REPLACE INTO app_settings VALUES ('eur_usd_rate',?)", (str(rate),))
                db.execute("INSERT OR REPLACE INTO app_settings VALUES ('eur_cache_time',?)", (str(_t2.time()),))
            return rate
    except:
        return stored

def to_usd(amount, currency, month_str=None):
    if currency == 'USD' or not currency: return amount
    if currency == 'ILS':
        rate = get_rate_for_month(month_str) if month_str else get_exchange_rate()
        return amount / rate
    if currency == 'EUR':
        return amount * get_eur_rate()
    return amount

def auto_migrate():
    """Add missing columns to existing tables without dropping data."""
    migrations = [
        ("expenses", "vendor_bank_num", "TEXT DEFAULT ''"),
        ("expenses", "vendor_branch_num", "TEXT DEFAULT ''"),
        ("expenses", "vendor_account_num", "TEXT DEFAULT ''"),
        ("expenses", "is_fixed", "INTEGER DEFAULT 1"),
        ("expenses", "payment_dates", "TEXT DEFAULT '{}'"),
        ("expenses", "expense_invoice_date", "TEXT DEFAULT ''"),
        ("expenses", "expense_payment_date", "TEXT DEFAULT ''"),
        ("expenses", "is_recurring", "INTEGER DEFAULT 1"),
        ("expenses", "subject_to_vat", "INTEGER DEFAULT 1"),
        ("contracts", "subject_to_vat", "INTEGER DEFAULT 0"),
        ("loans", "interest_start_month", "TEXT DEFAULT ''"),
        ("salary_snapshots", "net_pay_date", "TEXT DEFAULT ''"),
        ("salary_snapshots", "tax_pay_date", "TEXT DEFAULT ''"),
        ("salary_snapshots", "pension_pay_date", "TEXT DEFAULT ''"),
        ("expense_payments", "invoice_date", "TEXT DEFAULT ''"),
        ("employees", "pension_pct", "REAL DEFAULT 8.33"),
        ("employees", "disability_pct", "REAL DEFAULT 2.5"),
        ("employees", "study_fund_pct", "REAL DEFAULT 7.5"),
        ("employees", "study_fund_salary", "REAL DEFAULT 0"),
        ("employees", "pay_net_date", "TEXT DEFAULT ''"),
        ("employees", "pay_tax_date", "TEXT DEFAULT ''"),
        ("employees", "pay_pension_date", "TEXT DEFAULT ''"),
        ("expenses", "currency", "TEXT DEFAULT 'ILS'"),
        ("contracts", "is_new_client", "INTEGER DEFAULT 1"),
        ("contracts", "invoice_date", "TEXT DEFAULT ''"),
        ("contracts", "payment_splits", "TEXT DEFAULT '[]'"),
        ("contracts", "payment_method", "TEXT DEFAULT ''"),
        ("contracts", "contact_name", "TEXT DEFAULT ''"),
        ("contracts", "contact_phone", "TEXT DEFAULT ''"),
        ("contracts", "contact_email", "TEXT DEFAULT ''"),
        ("contracts", "contact_linkedin", "TEXT DEFAULT ''"),
        ("sales_targets", "avg_deal_target", "REAL DEFAULT 0"),
        ("sales_targets", "cycle_days_target", "INTEGER DEFAULT 0"),
        ("sales_targets", "close_rate_target", "REAL DEFAULT 0"),
        ("sales_targets", "new_clients_target", "INTEGER DEFAULT 0"),
    ]
    with get_db() as db:
        for table, column, col_type in migrations:
            try:
                db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
                print(f"  Migration: added {table}.{column}")
            except Exception:
                pass  # Column already exists
        
        # Recreate expense_payments without UNIQUE constraint (allows multiple invoices per month)
        try:
            has_unique = False
            tbl_info = db.execute("SELECT sql FROM sqlite_master WHERE name='expense_payments'").fetchone()
            if tbl_info and "UNIQUE" in (tbl_info["sql"] or ""):
                has_unique = True
            if has_unique:
                db.execute("ALTER TABLE expense_payments RENAME TO expense_payments_old")
                db.execute("""CREATE TABLE expense_payments (
                    id TEXT PRIMARY KEY, expense_id TEXT NOT NULL, month TEXT NOT NULL,
                    amount REAL DEFAULT 0, invoice_date TEXT DEFAULT '', payment_date TEXT DEFAULT '',
                    notes TEXT DEFAULT '', created_at TEXT)""")
                db.execute("""INSERT INTO expense_payments (id, expense_id, month, amount, invoice_date, payment_date, notes, created_at)
                             SELECT id, expense_id, month, amount, 
                                    COALESCE(invoice_date,''), COALESCE(payment_date,''), 
                                    COALESCE(notes,''), created_at 
                             FROM expense_payments_old""")
                db.execute("DROP TABLE expense_payments_old")
                print("  Migration: recreated expense_payments (removed UNIQUE)")
        except Exception as e:
            print(f"  Migration expense_payments: {e}")

def is_seeded():
    with get_db() as db: return db.execute("SELECT COUNT(*) c FROM employees").fetchone()["c"] > 0

def seed_db():
    with get_db() as db:
        # Exchange rate default
        db.execute("INSERT OR IGNORE INTO app_settings VALUES ('usd_ils_rate','3.12')")
        db.execute("INSERT OR IGNORE INTO app_settings VALUES ('vat_rate','18')")  # Israel VAT 2026
        # Default payroll payment dates (day of month)
        db.execute("INSERT OR IGNORE INTO app_settings VALUES ('payroll_net_day','9')")
        db.execute("INSERT OR IGNORE INTO app_settings VALUES ('payroll_net_month_offset','1')")  # 1 = next month
        db.execute("INSERT OR IGNORE INTO app_settings VALUES ('payroll_tax_day','15')")
        db.execute("INSERT OR IGNORE INTO app_settings VALUES ('payroll_tax_month_offset','1')")
        db.execute("INSERT OR IGNORE INTO app_settings VALUES ('payroll_pension_day','15')")
        db.execute("INSERT OR IGNORE INTO app_settings VALUES ('payroll_pension_month_offset','1')")
        # Users
        admin_hash = hash_pw("admin123")
        sales_hash = hash_pw("sales123")
        db.execute("INSERT OR IGNORE INTO users VALUES (?,?,?,?,?,?,?)", ("u-admin","admin",admin_hash,"admin","Admin Manager","G&A",1))
        db.execute("INSERT OR IGNORE INTO users VALUES (?,?,?,?,?,?,?)", ("u-eduardo","eduardo",sales_hash,"sales_manager","Eduardo Borotchin","S&M",1))
        db.execute("INSERT OR IGNORE INTO users VALUES (?,?,?,?,?,?,?)", ("u-elad","elad",sales_hash,"sales","Elad Lev","S&M",0))
        db.execute("INSERT OR IGNORE INTO users VALUES (?,?,?,?,?,?,?)", ("u-tiki","tiki",sales_hash,"sales","Tiki Tavero","S&M",0))

        # Employees
        emps = [
            ("e01","Etti Berger","G&A",16487.46,100,0,"2024-01-01",None),
            ("e02","Shai Grumet","R&D",11314.92,100,0,"2024-01-01",None),
            ("e03","Arye Laskin","R&D",9051.94,100,0,"2024-01-01",None),
            ("e04","Oren Chappo","R&D",11961.49,100,0,"2024-01-01",None),
            ("e05","Eduardo Borotchin","S&M",10345.07,100,0,"2024-01-01",None),
            ("e06","Nitai Driel","R&D",387.94,25,0,"2024-01-01",None),
            ("e07","Elad Lev","S&M",8082.09,100,0,"2024-01-01",None),
            ("e08","Tiki Tavero","S&M",6465.67,100,0,"2024-01-01",None),
            ("e09","Meirav Zetz","G&A",2020.52,50,0,"2024-01-01",None),
            ("e10","Yaniv Barkai","R&D",19431.60,100,0,"2024-01-01","2026-01-31"),
            ("e11","Elad Sheskin","R&D",0,0,0,"2024-01-01","2025-12-31"),
        ]
        db.executemany("INSERT OR IGNORE INTO employees (id,name,dept,gross,position_pct,bonus,hire_date,term_date) VALUES (?,?,?,?,?,?,?,?)", emps)

        now = datetime.now().isoformat()
        cons = [
            ("c01","Ministry of Defence","Indonesia","Government",2026,1,"2026-01-01",800000,60,"03-2026",0,0.75,None,"Demo",now,"u-eduardo","[]",now,now),
            ("c02","Ministry of Defence","Indonesia","Government",2026,1,"2026-01-01",150000,120,"05-2026",0,0.75,None,"Email Sent",now,"u-eduardo","[]",now,now),
            ("c03","Serbia","Serbia","Government",2025,10,"2025-10-01",90000,240,"05-2026",0,0.5,None,"Negotiation",now,"u-elad","[]",now,now),
            ("c04","Practical Cyber Academy","Singapore","Academy",2026,2,"2026-02-01",155000,240,"09-2026",0,0.75,None,"Demo",now,"u-tiki","[]",now,now),
            ("c06","Elta","Israel","Enterprise",2025,10,"2025-10-01",5000,180,"03-2026",0,1.0,3483.87,"Signed",now,"u-elad","[]",now,now),
            ("c07","Elta","Israel","Enterprise",2025,10,"2025-10-01",900000,270,"06-2026",0,0.5,None,"Negotiation",now,"u-eduardo","[]",now,now),
            ("c11","Schools","Israel","Academy",2025,10,"2025-10-01",15000,120,"01-2026",0,1.0,11231.94,"Signed",now,"u-tiki","[]",now,now),
            ("c13","Improvate","Israel","Academy",2026,5,"2026-05-01",150000,60,"06-2026",1,1.0,12108.06,"Signed",now,"u-eduardo","[]",now,now),
            ("c16","Bank of Israel","Israel","Government",2026,1,"2026-01-01",15000,45,"02-2026",0,1.0,None,"Signed",now,"u-elad","[]",now,now),
            ("c18","DSA","Cyprus","Government",2025,10,"2025-10-01",56000,210,"04-2026",0,0.5,35404.52,"Demo",now,"u-tiki","[]",now,now),
            ("c23","MAG","Nigeria","Enterprise",2026,1,"2026-01-01",30000,45,"02-2026",0,1.0,22800,"Signed",now,"u-eduardo","[]",now,now),
            ("c25","Technion","Israel","Academy",2026,1,"2026-01-01",42000,270,"09-2026",1,1.0,3870.97,"Negotiation",now,"u-tiki","[]",now,now),
            ("c28","EU Funding","Greece","Government",2026,1,"2026-01-01",300000,270,"09-2026",0,0.25,None,"Initial Contact",now,"u-elad","[]",now,now),
            ("c31","Abu Dhabi","UAE","Government",2026,1,"2026-01-01",300000,270,"09-2026",0,0.75,None,"Demo",now,"u-eduardo","[]",now,now),
            ("c39","Gabon","Gabon","Government",2026,3,"2026-03-01",450000,120,"06-2026",0,0.5,None,"Email Sent",now,"u-tiki","[]",now,now),
            ("c40","KSV/023","Kosovo","Government",2026,1,"2026-01-01",60000,60,"03-2026",0,1.0,None,"Signed",now,"u-elad","[]",now,now),
            ("c42","Migdal","Israel","Enterprise",2026,1,"2026-01-01",20322.58,45,"02-2026",0,1.0,None,"Negotiation",now,"u-eduardo","[]",now,now),
            ("c48","Ivory Coast","Ivory Coast","Government",2026,1,"2026-01-01",80000,60,"03-2026",0,0.75,None,"Demo",now,"u-tiki","[]",now,now),
        ]
        for c in cons:
            db.execute("INSERT OR IGNORE INTO contracts (id,client,country,industry,yr,sm,sd,val,dso,pm,monthly,chance,ap,stage,stage_updated_at,salesperson_id,notes,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", c)

        exps = [
            ("x01","R&D","Tera Sky","Cloud",1,carry_forward({"01-2026":2866.24,"02-2026":2500})),
            ("x02","G&A","Atidim","Rent",0,carry_forward({"01-2026":7006.37,"02-2026":6875})),
            ("x06","G&A","Misc.","Supplies",0,{m:300 for m in ML}),
            ("x07","R&D","Cellcom","Internet",1,carry_forward({"01-2026":254.78,"02-2026":257.99})),
            ("x10","Training","Misc.","Training",1,{m:20000 for m in ML}),
            ("x11","G&A","Misc.","Professional",0,{m:10000 for m in ML}),
            ("x12","R&D","Nir","Content",1,carry_forward({"01-2026":11146.50,"02-2026":11286.86})),
            ("x16","S&M","Misc.","Advertising",0,{m:1500 for m in ML}),
            ("x17","G&A","Misc.","Bank Fees",0,{m:800 for m in ML}),
            ("x18","R&D","Tzahi","Content",1,carry_forward({"02-2026":5643.43})),
        ]
        for eid,dept,vendor,sub,cogs,amounts in exps:
            db.execute("INSERT OR IGNORE INTO expenses (id,dept,vendor,sub_cat,is_cogs,amounts) VALUES (?,?,?,?,?,?)",
                       (eid,dept,vendor,sub,cogs,json.dumps(amounts)))

        # Seed default payment methods
        pms = [
            ("pm_mizrachi", "Bank Mizrachi", "mizrachi", 1, 1),
            ("pm_yahav", "Bank Yahav", "yahav", 1, 2),
            ("pm_cc_mizrachi", "Credit Card (Mizrachi)", "mizrachi", 1, 3),
        ]
        db.executemany("INSERT OR IGNORE INTO payment_methods (id, name, bank_group, is_active, sort_order) VALUES (?,?,?,?,?)", pms)

# ══════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════
class LoginReq(BaseModel):
    username: str; password: str

@app.post("/api/login")
def login(body: LoginReq):
    with get_db() as db:
        u = db.execute("SELECT * FROM users WHERE username=?", (body.username,)).fetchone()
    if not u or not verify_pw(body.password, u["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    token = jwt.encode({"sub": u["id"], "role": u["role"], "name": u["display_name"],
                        "canSeeAll": bool(u["can_see_all"]),
                        "exp": time.time() + 86400*7}, JWT_SECRET, algorithm="HS256")
    return {"token": token, "role": u["role"], "name": u["display_name"],
            "userId": u["id"], "canSeeAll": bool(u["can_see_all"])}

def get_current_user(authorization: str = Header(None)):
    if not authorization: raise HTTPException(401, "Not authenticated")
    token = authorization.replace("Bearer ", "")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        return payload
    except: raise HTTPException(401, "Invalid token")

def require_admin(user=Depends(get_current_user)):
    if user["role"] != "admin": raise HTTPException(403, "Admin only")
    return user

def require_manager(user=Depends(get_current_user)):
    if user["role"] not in ("admin", "sales_manager"): raise HTTPException(403, "Manager only")
    return user

class VisibilityUpdate(BaseModel):
    userId: str
    canSeeAll: bool

@app.put("/api/users/visibility")
def set_visibility(b: VisibilityUpdate, user=Depends(require_manager)):
    with get_db() as db:
        db.execute("UPDATE users SET can_see_all=? WHERE id=?", (int(b.canSeeAll), b.userId))
    return {"ok": True}

@app.get("/api/team")
def get_team(user=Depends(get_current_user)):
    """Get sales team members with their visibility settings."""
    if user["role"] not in ("admin", "sales_manager"):
        raise HTTPException(403)
    with get_db() as db:
        rows = db.execute("SELECT id, display_name, username, role, can_see_all FROM users WHERE dept='S&M'").fetchall()
    return [dict(r) for r in rows]

# ══════════════════════════════════════════
# DATA ACCESS
# ══════════════════════════════════════════
def load_employees():
    with get_db() as db: return [dict(r) for r in db.execute("SELECT * FROM employees WHERE is_ghost=0").fetchall()]

def load_contracts():
    with get_db() as db:
        rows = db.execute("""SELECT c.*, u.display_name as salesperson_name
                             FROM contracts c LEFT JOIN users u ON c.salesperson_id=u.id""").fetchall()
        return [dict(r) for r in rows]

def load_expenses():
    with get_db() as db:
        rows = db.execute("SELECT * FROM expenses").fetchall()
        return [dict(r) | {"amounts": json.loads(dict(r)["amounts"]), 
                           "payment_dates": json.loads(dict(r).get("payment_dates","{}") or "{}")} for r in rows]

def load_income_sources():
    with get_db() as db:
        rows = db.execute("SELECT * FROM income_sources").fetchall()
        return [dict(r) for r in rows]

def load_salespeople():
    with get_db() as db:
        return [dict(r) for r in db.execute("SELECT id,display_name,username,role,can_see_all FROM users WHERE role IN ('sales','sales_manager') OR dept='S&M'").fetchall()]

def log_audit(entity_type, entity_id, field, old_val, new_val, user_id):
    with get_db() as db:
        db.execute("INSERT INTO audit_log (entity_type,entity_id,field,old_value,new_value,changed_by,changed_at) VALUES (?,?,?,?,?,?,?)",
                   (entity_type, entity_id, field, str(old_val), str(new_val), user_id, datetime.now().isoformat()))

# ══════════════════════════════════════════
# FINANCIAL ENGINE (unchanged)
# ══════════════════════════════════════════
def parse_pm(pm):
    try: p=pm.split("-"); return int(p[0])-1 if int(p[1])==FY and 1<=int(p[0])<=12 else None
    except: return None

def shift_pm(pm, days):
    if not days: return pm
    try: p=pm.split("-"); d=date(int(p[1]),int(p[0]),15)+timedelta(days=days); return f"{d.month:02d}-{d.year}"
    except: return pm

def compute(delay=0, ghosts=0, gs=15000, w=True):
    employees=load_employees(); contracts=load_contracts(); expenses=load_expenses()
    vat_rate = get_vat_rate()  # 0.18 for Israel 2026
    for i in range(ghosts):
        employees.append({"id":f"g{i}","name":f"Ghost {i+1}","dept":"R&D","gross":gs,"hire_date":"2026-01-01","term_date":None})
    # Build invoice lookup: expense_id -> month -> total invoice amount
    inv_by_expense_month = {}
    try:
        with get_db() as db:
            inv_rows = db.execute("SELECT expense_id, amount, currency, due_date FROM invoices WHERE due_date IS NOT NULL").fetchall()
            for inv in inv_rows:
                if not inv["due_date"]: continue
                try:
                    dd = date.fromisoformat(inv["due_date"])
                    if dd.year == FY:
                        inv_ml = f"{dd.month:02d}-{dd.year}"
                        key = (inv["expense_id"], inv_ml)
                        if key not in inv_by_expense_month:
                            inv_by_expense_month[key] = {"amount": 0, "currency": inv["currency"] or "ILS"}
                        inv_by_expense_month[key]["amount"] += inv["amount"]
                except: pass
    except: pass

    # Load salary snapshots for all months
    snap_by_emp_month = {}
    try:
        with get_db() as db:
            snap_rows = db.execute("SELECT * FROM salary_snapshots").fetchall()
            for sr in snap_rows:
                snap_by_emp_month[(sr["employee_id"], sr["month"])] = dict(sr)
    except: pass

    # Load ALL expense payment records (invoices)
    exp_payments = {}
    exp_invoices_by_inv_month = {}  # P&L: amounts WITHOUT VAT
    exp_invoices_by_pay_month = {}  # CF: amounts WITH VAT when applicable
    try:
        with get_db() as db:
            ep_rows = db.execute("SELECT ep.*, e.dept, e.is_cogs, e.currency, e.subject_to_vat, e.service_desc FROM expense_payments ep LEFT JOIN expenses e ON ep.expense_id=e.id").fetchall()
            print(f"  [compute] Loaded {len(ep_rows)} expense_payment records")
            for ep_row in ep_rows:
                ep = dict(ep_row)
                amt = ep["amount"] or 0
                if amt <= 0: continue
                dept = ep["dept"] or "G&A"
                is_cogs = ep["is_cogs"] or 0
                cur = ep["currency"] or "ILS"
                has_vat = bool(ep.get("subject_to_vat", 1))
                sub_cat = (ep.get("service_desc") or "").strip() or "Other"
                amt_with_vat = amt * (1 + vat_rate) if has_vat else amt
                
                # P&L: by invoice_date month (without VAT)
                inv_ml = ep.get("month", "")
                if ep.get("invoice_date"):
                    try:
                        inv_d = date.fromisoformat(ep["invoice_date"])
                        inv_ml = f"{inv_d.month:02d}-{inv_d.year}"
                    except: pass
                if inv_ml:
                    exp_invoices_by_inv_month.setdefault(inv_ml, []).append({"amt": amt, "dept": dept, "is_cogs": is_cogs, "cur": cur, "ml": inv_ml, "cat": sub_cat})
                    print(f"  [compute] Invoice: {sub_cat} {amt} {cur} -> P&L {inv_ml}")
                
                # Cash Flow: by payment_date month (with VAT)
                if ep.get("payment_date"):
                    try:
                        pay_d = date.fromisoformat(ep["payment_date"])
                        pay_ml = f"{pay_d.month:02d}-{pay_d.year}"
                        exp_invoices_by_pay_month.setdefault(pay_ml, []).append({"amt": amt_with_vat, "cur": cur, "ml": pay_ml})
                    except: pass
                elif inv_ml:
                    exp_invoices_by_pay_month.setdefault(inv_ml, []).append({"amt": amt_with_vat, "cur": cur, "ml": inv_ml})
    except Exception as ex:
        print(f"  [compute] ERROR loading expense_payments: {ex}")
    except: pass

    data=[]
    payroll_components = []
    for mi,ml in enumerate(ML):
        p=ml.split("-"); md=date(int(p[1]),int(p[0]),1)
        lab={"R&D":0,"S&M":0,"G&A":0,"Training":0,"Finance":0}
        month_ss=0; month_pension=0; month_sf=0; month_dis=0; month_gross=0
        for e in employees:
            h=date.fromisoformat(e["hire_date"]); t=date.fromisoformat(e["term_date"]) if e["term_date"] else None
            if md>=h and (not t or md<=t):
                # Use snapshot if available for this employee+month, else current data
                snap = snap_by_emp_month.get((e["id"], ml))
                if snap:
                    eg = snap["gross"]
                    eb = snap.get("bonus", 0)
                    ep = snap.get("position_pct", 100)
                    epen = snap.get("pension_pct", 8.33)
                    edis = snap.get("disability_pct", 2.5)
                    esf = snap.get("study_fund_pct", 7.5)
                    esfs = snap.get("study_fund_salary", 0)
                    ecur = snap.get("currency", "ILS")
                    edept = snap.get("dept", e["dept"])
                else:
                    eg = e["gross"]
                    eb = e.get("bonus", 0)
                    ep = e.get("position_pct", 100)
                    epen = e.get("pension_pct", 8.33)
                    edis = e.get("disability_pct", 2.5)
                    esf = e.get("study_fund_pct", 7.5)
                    esfs = e.get("study_fund_salary", 0)
                    ecur = e.get("currency", "ILS")
                    edept = e["dept"]
                
                total_cost, breakdown = calc_employer_cost(eg, eb, ep, epen, edis, esf, esfs, ecur)
                lab[edept] += to_usd(total_cost, ecur, ml)
                month_gross += to_usd(breakdown["gross"] + breakdown["bonus"], ecur, ml)
                month_ss += to_usd(breakdown["socialSecurity"], ecur, ml)
                month_pension += to_usd(breakdown["pension"], ecur, ml)
                month_sf += to_usd(breakdown["studyFund"], ecur, ml)
                month_dis += to_usd(breakdown["disability"], ecur, ml)
        payroll_components.append({"ss": round(month_ss,2), "pension": round(month_pension,2),
                                   "studyFund": round(month_sf,2), "disability": round(month_dis,2),
                                   "grossNet": round(month_gross,2)})
        vd={"R&D":0,"S&M":0,"G&A":0,"Training":0,"Finance":0}; vc=vo=0
        vd_cats = {}  # category -> amount (for P&L line items)
        # P&L: vendor expenses based ONLY on saved invoices (expense_payments by invoice_date)
        for inv_entry in exp_invoices_by_inv_month.get(ml, []):
            a = to_usd(inv_entry["amt"], inv_entry["cur"], ml)
            vd[inv_entry["dept"]] += a
            cat = inv_entry.get("cat", "Other")
            vd_cats[cat] = vd_cats.get(cat, 0) + a
            if inv_entry["is_cogs"] == 1: vc += a
            else: vo += a
        
        # Cash Flow: vendor expenses based on expense_payments
        # For each expense+accrual_month, check where it's actually paid
        vd_cash=0
        cogs=lab["R&D"]+vc; opex=(sum(lab.values())-lab["R&D"])+vo
        data.append({"month":ml[:2],"lab":{k:round(v,2) for k,v in lab.items()},"vd":{k:round(v,2) for k,v in vd.items()},"vdCats":{k:round(v,2) for k,v in vd_cats.items()},"cogs":round(cogs,2),"opex":round(opex,2),"totalExp":round(cogs+opex,2),"revenue":0.0,"cashIn":0.0,"vendorCashOut":0.0})
    # Build vendor cash out per month based on expense_payments by payment_date
    vendor_cash_by_month = [0.0] * 12
    for pay_ml, entries in exp_invoices_by_pay_month.items():
        try:
            parts = pay_ml.split("-")
            mi_pay = int(parts[0]) - 1
            if 0 <= mi_pay < 12 and int(parts[1]) == FY:
                for entry in entries:
                    vendor_cash_by_month[mi_pay] += to_usd(entry["amt"], entry["cur"], pay_ml)
        except: pass
    for i in range(12):
        data[i]["vendorCashOut"] = round(vendor_cash_by_month[i], 2)
    
    # ── Revenue from income_payments (not contracts) ──
    inc_by_inv_month = {}  # P&L: by invoice_date month (without VAT)
    inc_by_pay_month = {}  # CF: by payment_date month (with VAT)
    rev_cats_by_month = {}  # category breakdown for P&L
    try:
        with get_db() as db:
            ip_rows = db.execute("""SELECT ip.*, inc.currency, inc.subject_to_vat, inc.category 
                                   FROM income_payments ip 
                                   LEFT JOIN income_sources inc ON ip.income_id=inc.id""").fetchall()
            for ip in ip_rows:
                ip = dict(ip)
                amt = ip.get("amount") or 0
                if amt <= 0: continue
                cur = ip.get("currency") or "ILS"
                has_vat = bool(ip.get("subject_to_vat", 0))
                cat = (ip.get("category") or "").strip() or "Revenue"
                amt_with_vat = amt * (1 + vat_rate) if has_vat else amt
                
                # P&L: by invoice_date
                inv_ml = ""
                if ip.get("invoice_date"):
                    try:
                        inv_d = date.fromisoformat(ip["invoice_date"])
                        inv_ml = f"{inv_d.month:02d}-{inv_d.year}"
                    except: pass
                if inv_ml:
                    inc_by_inv_month.setdefault(inv_ml, []).append({"amt": amt, "cur": cur, "cat": cat})
                
                # CF: by payment_date (with VAT)
                if ip.get("payment_date"):
                    try:
                        pay_d = date.fromisoformat(ip["payment_date"])
                        pay_ml = f"{pay_d.month:02d}-{pay_d.year}"
                        inc_by_pay_month.setdefault(pay_ml, []).append({"amt": amt_with_vat, "cur": cur})
                    except: pass
                elif inv_ml:
                    inc_by_pay_month.setdefault(inv_ml, []).append({"amt": amt_with_vat, "cur": cur})
    except Exception as ex:
        print(f"  [compute] ERROR loading income_payments: {ex}")
    
    # Apply revenue to data
    for mi in range(12):
        ml = ML[mi]
        rev_cats = {}
        for entry in inc_by_inv_month.get(ml, []):
            v = to_usd(entry["amt"], entry["cur"], ml)
            data[mi]["revenue"] += v
            rev_cats[entry["cat"]] = rev_cats.get(entry["cat"], 0) + v
        data[mi]["revCats"] = {k: round(v, 2) for k, v in rev_cats.items()}
    
    # Apply cash in from income
    for pay_ml, entries in inc_by_pay_month.items():
        try:
            parts = pay_ml.split("-")
            mi_pay = int(parts[0]) - 1
            if 0 <= mi_pay < 12 and int(parts[1]) == FY:
                for entry in entries:
                    data[mi_pay]["cashIn"] += to_usd(entry["amt"], entry["cur"], pay_ml)
        except: pass
    # ── Loans: interest → P&L Finance, payments → Cash Flow ──
    loans = load_loans()
    for ln in loans:
        sch = calc_loan_schedule(ln["amount"], ln["term_months"], ln["annual_rate"], ln["start_date"], ln["repayment_type"], ln.get("interest_start_month",""))
        cur = ln.get("currency", "ILS")
        for s in sch:
            try:
                pd2 = date.fromisoformat(s["date"])
                if pd2.year == FY and 1 <= pd2.month <= 12:
                    mi4 = pd2.month - 1
                    interest_usd = to_usd(s["interest"], cur, ML[mi4])
                    payment_usd = to_usd(s["payment"], cur, ML[mi4])
                    # Interest → P&L as Finance expense (vendor category, not labor)
                    data[mi4]["vd"]["Finance"] = data[mi4]["vd"].get("Finance", 0) + interest_usd
                    data[mi4].setdefault("vdCats", {})
                    data[mi4]["vdCats"]["Interest"] = data[mi4]["vdCats"].get("Interest", 0) + interest_usd
                    # Loan repayment tracking for cash flow
                    data[mi4].setdefault("loanInterest", 0)
                    data[mi4]["loanInterest"] += interest_usd
                    data[mi4].setdefault("loanPrincipal", 0)
                    data[mi4]["loanPrincipal"] += to_usd(s["principal"], cur, ML[mi4])
                    data[mi4].setdefault("loanPayment", 0)
                    data[mi4]["loanPayment"] += payment_usd
            except: pass
    # Recalculate COGS/OpEx after adding loan interest
    for mi5 in range(12):
        d5 = data[mi5]
        vc5 = sum(v2 for k2, v2 in d5["vd"].items() if k2 == "R&D")  # simplified
        d5["cogs"] = round(d5["lab"]["R&D"] + vc5, 2) if "R&D" in d5["lab"] else d5["cogs"]
        d5["opex"] = round(sum(d5["lab"].values()) - d5["lab"].get("R&D", 0) + sum(d5["vd"].values()) - vc5, 2)
        d5["totalExp"] = round(d5["cogs"] + d5["opex"], 2)

    cum=0
    # Load tax payments from app_settings
    tax_payments = {}
    loan_opening_bal = 0
    try:
        with get_db() as db:
            tax_rows = db.execute("SELECT key, value FROM app_settings WHERE key LIKE 'tax_%'").fetchall()
            for r in tax_rows:
                m = int(r["key"].replace("tax_",""))
                tax_payments[m] = float(r["value"])
            # Loan opening balance (manual)
            lob = db.execute("SELECT value FROM app_settings WHERE key='loan_opening_balance'").fetchone()
            if lob: loan_opening_bal = float(lob["value"])
    except: pass
    
    # ── AR: Accounts Receivable ──
    # AR = cumulative revenue recognized (by invoice_date) - cumulative cash received
    # This correctly captures: invoice issued → AR increases, payment received → AR decreases
    ar_by_month = [0.0] * 12
    cum_invoiced = 0; cum_received = 0
    for i in range(12):
        cum_invoiced += data[i]["revenue"]  # P&L revenue (by invoice_date)
        cum_received += data[i]["cashIn"]   # Cash Flow receipts (by payment date/DSO)
        ar_by_month[i] = round(max(0, cum_invoiced - cum_received), 2)
    
    # ── AP: Accounts Payable ──
    # AP from vendor invoices: exists from due_date until paid_date
    # For expenses without specific invoices: accrued monthly, "paid" same month (AP = 0)
    ap_by_month = [0.0] * 12
    try:
        with get_db() as db:
            inv_rows = db.execute("SELECT amount, currency, due_date, paid_amount, paid_date, status FROM invoices").fetchall()
        for inv in inv_rows:
            if not inv["due_date"]: continue
            try:
                dd = date.fromisoformat(inv["due_date"])
                if dd.year != FY: continue
                due_mi = dd.month - 1
                amt_usd = to_usd(inv["amount"], inv["currency"] or "ILS", ML[due_mi])
                paid_usd = to_usd(inv["paid_amount"] or 0, inv["currency"] or "ILS", ML[due_mi])
                remaining = amt_usd - paid_usd
                if remaining <= 0: continue
                # AP exists from due_date month until paid_date month (exclusive)
                # If not paid, AP persists through Dec
                paid_mi = 12  # default: not paid within FY
                if inv["paid_date"] and inv["status"] == "paid":
                    try:
                        pd3 = date.fromisoformat(inv["paid_date"])
                        if pd3.year == FY: paid_mi = pd3.month  # exclusive: AP cleared in payment month
                        elif pd3.year < FY: continue  # paid before FY, no AP
                    except: pass
                # AP balance exists from due month through month before payment
                for mi6 in range(due_mi, min(paid_mi, 12)):
                    ap_by_month[mi6] += remaining
            except: pass
    except: pass
    
    # ── Loan balances per month ──
    loan_bal_by_month = [0.0] * 12
    loan_remaining = loan_opening_bal
    # Build map of new loans by start month
    loan_additions = [0.0] * 12
    for ln in loans:
        try:
            lsd = date.fromisoformat(ln["start_date"])
            if lsd.year == FY and 1 <= lsd.month <= 12:
                loan_additions[lsd.month - 1] += to_usd(ln["amount"], ln.get("currency","ILS"), ML[lsd.month-1])
        except: pass
    # Calculate balance per month: add new loans, subtract principal repayments
    for i in range(12):
        loan_remaining += loan_additions[i]
        loan_remaining -= data[i].get("loanPrincipal", 0)
        loan_bal_by_month[i] = round(max(0, loan_remaining), 2)
    
    # Load payroll payment defaults
    payroll_defaults = {"net_day": 9, "net_offset": 1, "tax_day": 15, "tax_offset": 1, "pension_day": 15, "pension_offset": 1}
    try:
        with get_db() as db:
            for key, field in [("payroll_net_day","net_day"),("payroll_net_month_offset","net_offset"),
                               ("payroll_tax_day","tax_day"),("payroll_tax_month_offset","tax_offset"),
                               ("payroll_pension_day","pension_day"),("payroll_pension_month_offset","pension_offset")]:
                r = db.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
                if r: payroll_defaults[field] = int(r["value"])
    except: pass
    
    # Build payroll cash flow arrays: distribute by payment dates
    cf_net_pay = [0.0] * 12     # Net salary cash out per month
    cf_tax_pay = [0.0] * 12     # Tax cash out per month
    cf_pension_pay = [0.0] * 12 # Pension+SS+SF+Disability cash out per month
    
    for accrual_mi in range(12):
        accrual_ml = ML[accrual_mi]
        pc = payroll_components[accrual_mi]
        
        # Default targets (next month)
        net_target_mi = min(accrual_mi + payroll_defaults["net_offset"], 11)
        tax_target_mi = min(accrual_mi + payroll_defaults["tax_offset"], 11)
        pension_target_mi = min(accrual_mi + payroll_defaults["pension_offset"], 11)
        
        # Level 1: Snapshot-level overrides (Save Payroll Month)
        for eid_ml, snap in snap_by_emp_month.items():
            if eid_ml[1] == accrual_ml:
                if snap.get("net_pay_date"):
                    try:
                        npd = date.fromisoformat(snap["net_pay_date"])
                        if npd.year == FY: net_target_mi = npd.month - 1
                    except: pass
                if snap.get("tax_pay_date"):
                    try:
                        tpd = date.fromisoformat(snap["tax_pay_date"])
                        if tpd.year == FY: tax_target_mi = tpd.month - 1
                    except: pass
                if snap.get("pension_pay_date"):
                    try:
                        ppd = date.fromisoformat(snap["pension_pay_date"])
                        if ppd.year == FY: pension_target_mi = ppd.month - 1
                    except: pass
                break
        
        # Level 2: Per-employee payment dates override everything
        # Distribute each employee's components individually
        p2 = accrual_ml.split("-")
        md2 = date(int(p2[1]), int(p2[0]), 1)
        emp_net = 0; emp_tax = 0; emp_pension = 0
        for e in employees:
            h2 = date.fromisoformat(e["hire_date"]); t2 = date.fromisoformat(e["term_date"]) if e["term_date"] else None
            if not (md2 >= h2 and (not t2 or md2 <= t2)): continue
            
            snap = snap_by_emp_month.get((e["id"], accrual_ml))
            eg = snap["gross"] if snap else e["gross"]
            eb = snap.get("bonus",0) if snap else e.get("bonus",0)
            ep = snap.get("position_pct",100) if snap else e.get("position_pct",100)
            ecur = snap.get("currency","ILS") if snap else e.get("currency","ILS")
            epen = snap.get("pension_pct",8.33) if snap else e.get("pension_pct",8.33)
            edis = snap.get("disability_pct",2.5) if snap else e.get("disability_pct",2.5)
            esf = snap.get("study_fund_pct",7.5) if snap else e.get("study_fund_pct",7.5)
            esfs = snap.get("study_fund_salary",0) if snap else e.get("study_fund_salary",0)
            
            _, bd = calc_employer_cost(eg, eb, ep, epen, edis, esf, esfs, ecur)
            
            # Per-employee payment date overrides
            e_net_mi = net_target_mi
            e_tax_mi = tax_target_mi
            e_pen_mi = pension_target_mi
            
            if e.get("pay_net_date"):
                try:
                    nd = date.fromisoformat(e["pay_net_date"])
                    offset = (nd.year - FY) * 12 + nd.month - 1 - 0
                    e_net_mi = min(max(accrual_mi + offset, 0), 11) if offset != 0 else nd.month - 1
                    if nd.year == FY: e_net_mi = nd.month - 1
                except: pass
            if e.get("pay_tax_date"):
                try:
                    td2 = date.fromisoformat(e["pay_tax_date"])
                    if td2.year == FY: e_tax_mi = td2.month - 1
                except: pass
            if e.get("pay_pension_date"):
                try:
                    pd4 = date.fromisoformat(e["pay_pension_date"])
                    if pd4.year == FY: e_pen_mi = pd4.month - 1
                except: pass
            
            cf_net_pay[e_net_mi] += to_usd(bd["gross"] + bd["bonus"], ecur, accrual_ml)
            cf_pension_pay[e_pen_mi] += to_usd(bd["socialSecurity"] + bd["pension"] + bd["studyFund"] + bd["disability"], ecur, accrual_ml)
        
        cf_tax_pay[tax_target_mi] += tax_payments.get(accrual_mi + 1, 0)
    
    for i, d in enumerate(data):
        d["ar"] = ar_by_month[i]
        d["ap"] = round(ap_by_month[i], 2)
        d["loanBalance"] = loan_bal_by_month[i]
        d["revenue"]=round(d["revenue"],2); d["cashIn"]=round(d["cashIn"],2)
        d["grossProfit"]=round(d["revenue"]-d["cogs"],2)
        d["grossMargin"]=round(d["grossProfit"]/d["revenue"]*100,1) if d["revenue"]>0 else 0
        d["netIncome"]=round(d["revenue"]-d["totalExp"],2)
        
        # Payroll CF from pre-computed arrays
        d["cfGross"] = round(cf_net_pay[i], 2)
        d["cfSS"] = 0  # included in cfPension below
        d["cfPension"] = round(cf_pension_pay[i], 2)
        d["cfStudyFund"] = 0  # included in cfPension
        d["cfDisability"] = 0  # included in cfPension
        d["cfTax"] = round(cf_tax_pay[i], 2)
        
        # Payroll accruals for balance sheet (what we owe at end of month)
        pc = payroll_components[i]
        d["accrualGross"] = round(pc["grossNet"], 2)
        d["accrualSS"] = round(pc["ss"], 2)
        d["accrualPension"] = round(pc["pension"], 2)
        d["accrualSF"] = round(pc["studyFund"], 2)
        d["accrualDisability"] = round(pc["disability"], 2)
        d["accrualTax"] = round(tax_payments.get(i + 1, 0), 2)
        
        # Total payroll cash out
        d["cfPayrollTotal"] = round(d["cfGross"] + d["cfPension"] + d["cfTax"], 2)
        
        d["cashOut"]=round(d["cfPayrollTotal"] + d.get("vendorCashOut", sum(d["vd"].values())), 2)
        d["netCashflow"]=round(d["cashIn"]-d["cashOut"],2)
        cum+=d["netCashflow"]; d["cumCash"]=round(cum,2)
        
        # Burn rates — based on CASH FLOW, not P&L
        # Net burn = OpEx cash out (no COGS) minus cash in
        opex_cash = sum(d["lab"].values()) - d["lab"].get("R&D", 0) + sum(d["vd"].values()) - sum(v for k,v in d["vd"].items() if k == "R&D")
        d["burnNet"] = round(opex_cash - d["cashIn"], 2) if opex_cash > d["cashIn"] else 0
        # Gross burn = Total cash out (including loan payments) minus cash in
        total_cash_out = d["cashOut"] + d.get("loanPayment", 0)
        d["burnGross"] = round(total_cash_out - d["cashIn"], 2) if total_cash_out > d["cashIn"] else 0
        # Legacy
        d["burnRate"] = d["burnGross"]
        # Runway based on gross burn
        d["runway"] = round(cum / d["burnGross"], 1) if d["burnGross"] > 0 else None
    
    # Quarterly aggregation — based on Cash Flow
    for qi in range(4):
        ms = data[qi*3:qi*3+3]
        if not ms: continue
        q_rev = sum(m["revenue"] for m in ms)
        q_cashIn = sum(m["cashIn"] for m in ms)
        q_cashOut = sum(m["cashOut"] + m.get("loanPayment", 0) for m in ms)
        q_opexCash = sum(sum(m["lab"].values()) - m["lab"].get("R&D",0) + sum(m["vd"].values()) - sum(v for k,v in m["vd"].items() if k=="R&D") for m in ms)
        q_burnNet = round(q_opexCash - q_cashIn, 2) if q_opexCash > q_cashIn else 0
        q_burnGross = round(q_cashOut - q_cashIn, 2) if q_cashOut > q_cashIn else 0
        q_label = f"Q{qi+1}"
        for m in ms:
            m["quarter"] = q_label
            m["qBurnNet"] = q_burnNet
            m["qBurnGross"] = q_burnGross
            m["qRevenue"] = round(q_rev, 2)
            m["qExpenses"] = round(q_cashOut, 2)
            m["qNetCashflow"] = round(q_cashIn - q_cashOut, 2)
    
    # Runway recalc with trailing 3-month avg gross burn
    for i, d in enumerate(data):
        start = max(0, i - 2)
        trailing = data[start:i+1]
        avg_burn = sum(m["burnGross"] for m in trailing) / len(trailing)
        d["runwayAvg"] = round(d["cumCash"] / avg_burn, 1) if avg_burn > 0 else None
    
    return data

# ══════════════════════════════════════════
# API: Data
# ══════════════════════════════════════════
@app.get("/api/data")
def get_data(weighted:bool=True, delay:int=0, ghosts:int=0, salary:float=15000):
    s=compute(delay,ghosts,salary,weighted)
    contracts=load_contracts(); employees=load_employees(); expenses=load_expenses()
    cs=[{"id":c["id"],"client":c["client"],"country":c["country"],"industry":c["industry"],
         "value":c["val"],"chance":c["chance"],"weighted":round(c["val"]*(c["chance"] or 0),2),
         "dso":c["dso"],"pm":c["pm"],"monthly":bool(c["monthly"]),"signingDate":c["sd"],
         "currency":c.get("currency","USD"),
         "stage":c["stage"],"stageUpdatedAt":c["stage_updated_at"],
         "salespersonId":c["salesperson_id"],"salespersonName":c.get("salesperson_name",""),
         "notes":json.loads(c["notes"]) if c["notes"] else [],
         "contactName":c.get("contact_name",""),"contactPhone":c.get("contact_phone",""),
         "contactEmail":c.get("contact_email",""),"contactLinkedin":c.get("contact_linkedin",""),
         "paymentSplits":json.loads(c.get("payment_splits","[]")) if c.get("payment_splits") else [],
         "paymentMethod":c.get("payment_method",""),
         "isNewClient":bool(c.get("is_new_client",1)),
         "invoiceDate":c.get("invoice_date",""),
         "subjectToVat":bool(c.get("subject_to_vat",0)),
         "createdAt":c["created_at"],"updatedAt":c["updated_at"]} for c in contracts]
    es=[]
    for e in employees:
        total_cost, breakdown = calc_employer_cost(
            e["gross"], e.get("bonus",0), e.get("position_pct",100),
            e.get("pension_pct",8.33), e.get("disability_pct",2.5),
            e.get("study_fund_pct",7.5), e.get("study_fund_salary",0),
            e.get("currency","ILS"))
        es.append({"id":e["id"],"name":e["name"],"dept":e["dept"],"gross":e["gross"],
         "positionPct":e.get("position_pct",100),"bonus":e.get("bonus",0),
         "currency":e.get("currency","ILS"),
         "pensionPct":e.get("pension_pct",8.33),
         "disabilityPct":e.get("disability_pct",2.5),
         "studyFundPct":e.get("study_fund_pct",7.5),
         "studyFundSalary":e.get("study_fund_salary",0),
         "costBreakdown":breakdown,
         "laborCost":round(to_usd(total_cost, e.get("currency","ILS")),2),
         "laborCostMonthly":round(total_cost,2),
         "hireDate":e["hire_date"],"termDate":e["term_date"],
         "payNetDate":e.get("pay_net_date",""),"payTaxDate":e.get("pay_tax_date",""),"payPensionDate":e.get("pay_pension_date","")})
    xs=[{"id":x["id"],"dept":x["dept"],"vendor":x["vendor"],"subCat":x["sub_cat"],
         "isCogs":x["is_cogs"]==1,"expenseType":["OpEx","COGS","CapEx"][min(x["is_cogs"],2)],"currency":x.get("currency","ILS"),
         "amount":round(to_usd(sum(x["amounts"].values())/max(len(x["amounts"]),1), x.get("currency","ILS")),2),
         "amountOriginal":round(sum(x["amounts"].values())/max(len(x["amounts"]),1),2),
         "vendorContact":x.get("vendor_contact",""),"vendorEmail":x.get("vendor_email",""),
         "vendorPhone":x.get("vendor_phone",""),"serviceDesc":x.get("service_desc",""),
         "vendorBankNum":x.get("vendor_bank_num",""),"vendorBranchNum":x.get("vendor_branch_num",""),
         "vendorAccountNum":x.get("vendor_account_num",""),
         "isFixed":bool(x.get("is_fixed",1)),
         "paymentDates":x.get("payment_dates",{}),
         "expenseInvoiceDate":x.get("expense_invoice_date",""),
         "expensePaymentDate":x.get("expense_payment_date",""),
         "isRecurring":bool(x.get("is_recurring",1)),
         "subjectToVat":bool(x.get("subject_to_vat",1)),
         "frequency":x.get("frequency","monthly"),
         "paymentMethod":x.get("payment_method","")} for x in expenses]
    sp=load_salespeople()
    rate=get_exchange_rate()
    # Load payment methods
    with get_db() as db:
        pm_rows = db.execute("SELECT * FROM payment_methods WHERE is_active=1 ORDER BY sort_order").fetchall()
    pms = [{"id":p["id"],"name":p["name"],"bankGroup":p["bank_group"]} for p in pm_rows]
    return {"summary":s,"contracts":cs,"employees":es,"expenses":xs,
            "vatRate": round(get_vat_rate() * 100),
            "salespeople":[{"id":p["id"],"name":p["display_name"]} for p in sp],
            "stages":STAGES,"exchangeRate":rate,"paymentMethods":pms,
            "loans":[{"id":ln["id"],"lender":ln["lender"],"amount":ln["amount"],"termMonths":ln["term_months"],
                      "annualRate":ln["annual_rate"],"startDate":ln["start_date"],"repaymentType":ln["repayment_type"],
                      "currency":ln.get("currency","ILS"),"notes":ln.get("notes",""),
                      "interestStartMonth":ln.get("interest_start_month","")} for ln in load_loans()],
            "incomeSources": [{"id":inc["id"],"client":inc["client"],"category":inc.get("category",""),
                               "currency":inc.get("currency","ILS"),
                               "contactName":inc.get("contact_name",""),"contactPhone":inc.get("contact_phone",""),
                               "contactEmail":inc.get("contact_email",""),
                               "isRecurring":bool(inc.get("is_recurring",0)),
                               "subjectToVat":bool(inc.get("subject_to_vat",0)),
                               "notes":inc.get("notes","")} for inc in load_income_sources()]}

@app.get("/api/sales-analytics")
def sales_analytics(weighted:bool=True):
    contracts=load_contracts()
    by_type={}; by_country={}; total_signed=0; total_pipeline=0
    for c in contracts:
        v=c["val"]
        if weighted: v=v*c["chance"] if c["chance"] is not None else 0
        ind=c["industry"] or "Other"; co=c["country"] or "Unknown"
        is_signed = c["stage"] == "Signed"

        by_type.setdefault(ind,{"signed":0,"pipeline":0,"signedCount":0,"pipelineCount":0})
        by_country.setdefault(co,{"signed":0,"pipeline":0,"signedCount":0,"pipelineCount":0})

        if is_signed:
            by_type[ind]["signed"]+=v; by_type[ind]["signedCount"]+=1
            by_country[co]["signed"]+=v; by_country[co]["signedCount"]+=1
            total_signed+=v
        else:
            by_type[ind]["pipeline"]+=v; by_type[ind]["pipelineCount"]+=1
            by_country[co]["pipeline"]+=v; by_country[co]["pipelineCount"]+=1
            total_pipeline+=v

    total = total_signed + total_pipeline
    for k in by_type:
        d=by_type[k]; d["total"]=round(d["signed"]+d["pipeline"],2)
        d["signed"]=round(d["signed"],2); d["pipeline"]=round(d["pipeline"],2)
        d["pct"]=round(d["total"]/total*100,1) if total>0 else 0
        d["count"]=d["signedCount"]+d["pipelineCount"]
    for k in by_country:
        d=by_country[k]; d["total"]=round(d["signed"]+d["pipeline"],2)
        d["signed"]=round(d["signed"],2); d["pipeline"]=round(d["pipeline"],2)
        d["pct"]=round(d["total"]/total*100,1) if total>0 else 0
        d["count"]=d["signedCount"]+d["pipelineCount"]

    # Top = by signed only (real revenue)
    top_type=max(by_type,key=lambda k:by_type[k]["signed"]) if by_type else None
    top_country=max(by_country,key=lambda k:by_country[k]["signed"]) if by_country else None
    stages_dist={s:0 for s in STAGES}
    for c in contracts: stages_dist[c["stage"]]=stages_dist.get(c["stage"],0)+1
    return {
        "totalSigned":round(total_signed,2),"totalPipeline":round(total_pipeline,2),
        "total":round(total,2),
        "byType":by_type,
        "byCountry":dict(sorted(by_country.items(),key=lambda x:-x[1]["total"])[:10]),
        "topType":top_type,"topCountry":top_country,"dealCount":len(contracts),
        "weighted":weighted,"stageDistribution":stages_dist
    }

# ══════════════════════════════════════════
# API: CRUD (auth required for writes)
# ══════════════════════════════════════════
class ContractIn(BaseModel):
    client:str; country:str=""; industry:str="Enterprise"; value:float
    signingDate:str; dso:int=60; monthly:bool=False; chance:Optional[float]=None
    salespersonId:Optional[str]=None; currency:str="USD"
    contactName:str=""; contactPhone:str=""; contactEmail:str=""; contactLinkedin:str=""

@app.post("/api/contracts")
def add_contract(b:ContractIn, user=Depends(get_current_user)):
    cid=str(uuid4())[:8]; sd=date.fromisoformat(b.signingDate)
    pd=sd+timedelta(days=b.dso); pm=f"{pd.month:02d}-{pd.year}"
    now=datetime.now().isoformat()
    sp_id = b.salespersonId or (user["sub"] if user["role"] in ("sales","sales_manager") else None)
    with get_db() as db:
        db.execute("INSERT INTO contracts (id,client,country,industry,yr,sm,sd,val,dso,pm,monthly,chance,currency,stage,stage_updated_at,salesperson_id,notes,contact_name,contact_phone,contact_email,contact_linkedin,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                   (cid,b.client,b.country,b.industry,sd.year,sd.month,b.signingDate,b.value,b.dso,pm,int(b.monthly),b.chance,b.currency,"Initial Contact",now,sp_id,"[]",b.contactName,b.contactPhone,b.contactEmail,b.contactLinkedin,now,now))
    return {"ok":True,"id":cid,"paymentMonth":pm}

class StageUpdate(BaseModel):
    stage: str

@app.put("/api/contracts/{cid}/stage")
def update_stage(cid:str, b:StageUpdate, user=Depends(get_current_user)):
    if b.stage not in STAGES: raise HTTPException(400, "Invalid stage")
    with get_db() as db:
        c = db.execute("SELECT stage, salesperson_id FROM contracts WHERE id=?", (cid,)).fetchone()
        if not c: raise HTTPException(404)
        # Sales can only edit their own
        if user["role"] == "sales" and c["salesperson_id"] != user["sub"]:
            raise HTTPException(403, "Can only edit your own deals")
        old_idx = STAGES.index(c["stage"]) if c["stage"] in STAGES else 0
        new_idx = STAGES.index(b.stage)
        if new_idx < old_idx: raise HTTPException(400, "Cannot move stage backwards")
        now = datetime.now().isoformat()
        log_audit("contract", cid, "stage", c["stage"], b.stage, user["sub"])
        db.execute("UPDATE contracts SET stage=?, stage_updated_at=?, updated_at=? WHERE id=?", (b.stage, now, now, cid))
    return {"ok": True}

class NoteIn(BaseModel):
    text: str

@app.post("/api/contracts/{cid}/notes")
def add_note(cid:str, b:NoteIn, user=Depends(get_current_user)):
    now = datetime.now().isoformat()
    with get_db() as db:
        c = db.execute("SELECT notes, salesperson_id FROM contracts WHERE id=?", (cid,)).fetchone()
        if not c: raise HTTPException(404)
        if user["role"] == "sales" and c["salesperson_id"] != user["sub"]:
            raise HTTPException(403, "Can only add notes to your own deals")
        notes = json.loads(c["notes"]) if c["notes"] else []
        notes.append({"text": b.text, "by": user.get("name", user["sub"]), "at": now})
        db.execute("UPDATE contracts SET notes=?, updated_at=? WHERE id=?", (json.dumps(notes), now, cid))
    return {"ok": True}

class SplitsUpdate(BaseModel):
    splits: list  # [{amount: float, date: "YYYY-MM-DD"}, ...]

@app.put("/api/contracts/{cid}/splits")
def update_splits(cid:str, b:SplitsUpdate, user=Depends(get_current_user)):
    now = datetime.now().isoformat()
    with get_db() as db:
        c = db.execute("SELECT salesperson_id FROM contracts WHERE id=?", (cid,)).fetchone()
        if not c: raise HTTPException(404)
        if user["role"] == "sales" and c["salesperson_id"] != user["sub"]:
            raise HTTPException(403)
        db.execute("UPDATE contracts SET payment_splits=?, updated_at=? WHERE id=?",
                   (json.dumps(b.splits), now, cid))
    return {"ok": True}

class ContractUpdate(BaseModel):
    client:Optional[str]=None; country:Optional[str]=None; industry:Optional[str]=None
    value:Optional[float]=None; dso:Optional[int]=None; chance:Optional[float]=None
    monthly:Optional[bool]=None; salespersonId:Optional[str]=None
    contactName:Optional[str]=None; contactPhone:Optional[str]=None
    contactEmail:Optional[str]=None; contactLinkedin:Optional[str]=None
    paymentMethod:Optional[str]=None; isNewClient:Optional[bool]=None
    invoiceDate:Optional[str]=None
    subjectToVat:Optional[bool]=None

@app.put("/api/contracts/{cid}")
def upd_contract(cid:str, b:ContractUpdate, user=Depends(get_current_user)):
    now = datetime.now().isoformat()
    fields=[]; vals=[]
    if b.client is not None: fields.append("client=?"); vals.append(b.client)
    if b.country is not None: fields.append("country=?"); vals.append(b.country)
    if b.industry is not None: fields.append("industry=?"); vals.append(b.industry)
    if b.value is not None: fields.append("val=?"); vals.append(b.value)
    if b.dso is not None: fields.append("dso=?"); vals.append(b.dso)
    if b.chance is not None: fields.append("chance=?"); vals.append(b.chance)
    if b.monthly is not None: fields.append("monthly=?"); vals.append(int(b.monthly))
    if b.salespersonId is not None:
        # Only admin and sales_manager can reassign salesperson
        if user["role"] in ("admin", "sales_manager"):
            fields.append("salesperson_id=?"); vals.append(b.salespersonId)
    if b.contactName is not None: fields.append("contact_name=?"); vals.append(b.contactName)
    if b.contactPhone is not None: fields.append("contact_phone=?"); vals.append(b.contactPhone)
    if b.contactEmail is not None: fields.append("contact_email=?"); vals.append(b.contactEmail)
    if b.contactLinkedin is not None: fields.append("contact_linkedin=?"); vals.append(b.contactLinkedin)
    if b.paymentMethod is not None: fields.append("payment_method=?"); vals.append(b.paymentMethod)
    if b.isNewClient is not None: fields.append("is_new_client=?"); vals.append(int(b.isNewClient))
    if b.invoiceDate is not None: fields.append("invoice_date=?"); vals.append(b.invoiceDate)
    if b.subjectToVat is not None: fields.append("subject_to_vat=?"); vals.append(int(b.subjectToVat))
    if not fields: return {"ok":True}
    fields.append("updated_at=?"); vals.append(now)
    vals.append(cid)
    with get_db() as db:
        # Sales can only edit own
        if user["role"] == "sales":
            c = db.execute("SELECT salesperson_id FROM contracts WHERE id=?", (cid,)).fetchone()
            if c and c["salesperson_id"] != user["sub"]: raise HTTPException(403)
        db.execute(f"UPDATE contracts SET {','.join(fields)} WHERE id=?", vals)
        if b.dso is not None:
            row = db.execute("SELECT sd,dso FROM contracts WHERE id=?", (cid,)).fetchone()
            if row:
                sd=date.fromisoformat(row["sd"]); pd=sd+timedelta(days=row["dso"])
                db.execute("UPDATE contracts SET pm=? WHERE id=?", (f"{pd.month:02d}-{pd.year}", cid))
    return {"ok":True}

@app.delete("/api/contracts/{cid}")
def del_contract(cid:str, user=Depends(require_admin)):
    with get_db() as db: db.execute("DELETE FROM contracts WHERE id=?", (cid,))
    return {"ok":True}

class EmployeeIn(BaseModel):
    name:str; department:str="R&D"; gross:float; hireDate:str; termDate:Optional[str]=None

@app.post("/api/employees")
def add_employee(b:EmployeeIn, user=Depends(require_admin)):
    eid=str(uuid4())[:8]
    with get_db() as db:
        db.execute("INSERT INTO employees (id,name,dept,gross,hire_date,term_date) VALUES (?,?,?,?,?,?)",
                   (eid,b.name,b.department,b.gross,b.hireDate,b.termDate))
    return {"ok":True,"id":eid,"laborCost":round(b.gross*ECF,2)}

class EmployeeUpdate(BaseModel):
    name:Optional[str]=None; department:Optional[str]=None
    gross:Optional[float]=None; positionPct:Optional[float]=None
    bonus:Optional[float]=None; currency:Optional[str]=None; termDate:Optional[str]=None
    pensionPct:Optional[float]=None; disabilityPct:Optional[float]=None
    studyFundPct:Optional[float]=None; studyFundSalary:Optional[float]=None
    payNetDate:Optional[str]=None; payTaxDate:Optional[str]=None; payPensionDate:Optional[str]=None

@app.put("/api/employees/{eid}")
def upd_employee(eid:str, b:EmployeeUpdate, user=Depends(require_admin)):
    fields=[]; vals=[]
    if b.name is not None: fields.append("name=?"); vals.append(b.name)
    if b.department is not None: fields.append("dept=?"); vals.append(b.department)
    if b.gross is not None: fields.append("gross=?"); vals.append(b.gross)
    if b.positionPct is not None: fields.append("position_pct=?"); vals.append(b.positionPct)
    if b.bonus is not None: fields.append("bonus=?"); vals.append(b.bonus)
    if b.currency is not None: fields.append("currency=?"); vals.append(b.currency)
    if b.termDate is not None: fields.append("term_date=?"); vals.append(b.termDate if b.termDate else None)
    if b.pensionPct is not None: fields.append("pension_pct=?"); vals.append(b.pensionPct)
    if b.disabilityPct is not None: fields.append("disability_pct=?"); vals.append(b.disabilityPct)
    if b.studyFundPct is not None: fields.append("study_fund_pct=?"); vals.append(b.studyFundPct)
    if b.studyFundSalary is not None: fields.append("study_fund_salary=?"); vals.append(b.studyFundSalary)
    if b.payNetDate is not None: fields.append("pay_net_date=?"); vals.append(b.payNetDate)
    if b.payTaxDate is not None: fields.append("pay_tax_date=?"); vals.append(b.payTaxDate)
    if b.payPensionDate is not None: fields.append("pay_pension_date=?"); vals.append(b.payPensionDate)
    if not fields: return {"ok":True}
    vals.append(eid)
    with get_db() as db:
        db.execute(f"UPDATE employees SET {','.join(fields)} WHERE id=?", vals)
    return {"ok":True}

@app.delete("/api/employees/{eid}")
def del_employee(eid:str, user=Depends(require_admin)):
    with get_db() as db: db.execute("DELETE FROM employees WHERE id=?", (eid,))
    return {"ok":True}

class ExpenseIn(BaseModel):
    vendor:str; department:str="G&A"; subCategory:str=""; monthlyAmount:float=0; isCogs:int=0
    vendorContact:str=""; vendorEmail:str=""; vendorPhone:str=""; serviceDesc:str=""
    frequency:str="monthly"; currency:str="ILS"

FREQ_MULTIPLIERS = {"monthly":1,"one_time":0,"bimonthly":0.5,"quarterly":1/3,"semi_annual":1/6,"annual":1/12}

@app.post("/api/expenses")
def add_expense(b:ExpenseIn, user=Depends(require_admin)):
    xid=str(uuid4())[:8]
    mult=FREQ_MULTIPLIERS.get(b.frequency, 1)
    amounts={m: round(b.monthlyAmount * mult, 2) for m in ML}
    with get_db() as db:
        db.execute("INSERT INTO expenses (id,dept,vendor,sub_cat,is_cogs,amounts,currency,vendor_contact,vendor_email,vendor_phone,service_desc,frequency) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                   (xid,b.department,b.vendor,b.subCategory,int(b.isCogs),json.dumps(amounts),b.currency,b.vendorContact,b.vendorEmail,b.vendorPhone,b.serviceDesc,b.frequency))
    return {"ok":True,"id":xid}

class ExpenseUpdate(BaseModel):
    vendor:Optional[str]=None; department:Optional[str]=None; subCategory:Optional[str]=None
    monthlyAmount:Optional[float]=None; isCogs:Optional[int]=None
    vendorContact:Optional[str]=None; vendorEmail:Optional[str]=None
    vendorPhone:Optional[str]=None; serviceDesc:Optional[str]=None
    frequency:Optional[str]=None; currency:Optional[str]=None
    paymentMethod:Optional[str]=None
    vendorBankNum:Optional[str]=None; vendorBranchNum:Optional[str]=None
    vendorAccountNum:Optional[str]=None; isFixed:Optional[bool]=None
    paymentDates:Optional[dict]=None
    expenseInvoiceDate:Optional[str]=None; expensePaymentDate:Optional[str]=None
    isRecurring:Optional[bool]=None
    subjectToVat:Optional[bool]=None

@app.put("/api/expenses/{xid}")
def upd_expense(xid:str, b:ExpenseUpdate, user=Depends(require_admin)):
    fields=[]; vals=[]
    if b.vendor is not None: fields.append("vendor=?"); vals.append(b.vendor)
    if b.department is not None: fields.append("dept=?"); vals.append(b.department)
    if b.subCategory is not None: fields.append("sub_cat=?"); vals.append(b.subCategory)
    if b.isCogs is not None: fields.append("is_cogs=?"); vals.append(int(b.isCogs))
    if b.vendorContact is not None: fields.append("vendor_contact=?"); vals.append(b.vendorContact)
    if b.vendorEmail is not None: fields.append("vendor_email=?"); vals.append(b.vendorEmail)
    if b.vendorPhone is not None: fields.append("vendor_phone=?"); vals.append(b.vendorPhone)
    if b.serviceDesc is not None: fields.append("service_desc=?"); vals.append(b.serviceDesc)
    if b.frequency is not None: fields.append("frequency=?"); vals.append(b.frequency)
    if b.currency is not None: fields.append("currency=?"); vals.append(b.currency)
    if b.paymentMethod is not None: fields.append("payment_method=?"); vals.append(b.paymentMethod)
    if b.vendorBankNum is not None: fields.append("vendor_bank_num=?"); vals.append(b.vendorBankNum)
    if b.vendorBranchNum is not None: fields.append("vendor_branch_num=?"); vals.append(b.vendorBranchNum)
    if b.vendorAccountNum is not None: fields.append("vendor_account_num=?"); vals.append(b.vendorAccountNum)
    if b.isFixed is not None: fields.append("is_fixed=?"); vals.append(int(b.isFixed))
    if b.paymentDates is not None: fields.append("payment_dates=?"); vals.append(json.dumps(b.paymentDates))
    if b.expenseInvoiceDate is not None: fields.append("expense_invoice_date=?"); vals.append(b.expenseInvoiceDate)
    if b.expensePaymentDate is not None: fields.append("expense_payment_date=?"); vals.append(b.expensePaymentDate)
    if b.isRecurring is not None: fields.append("is_recurring=?"); vals.append(int(b.isRecurring))
    if b.subjectToVat is not None: fields.append("subject_to_vat=?"); vals.append(int(b.subjectToVat))
    if b.monthlyAmount is not None or b.frequency is not None:
        freq = b.frequency
        amt = b.monthlyAmount
        if freq is None or amt is None:
            with get_db() as db:
                row = db.execute("SELECT amounts, frequency FROM expenses WHERE id=?", (xid,)).fetchone()
                if row:
                    if freq is None: freq = row["frequency"]
                    if amt is None:
                        old_amounts = json.loads(row["amounts"])
                        avg = sum(old_amounts.values()) / max(len(old_amounts), 1)
                        old_mult = FREQ_MULTIPLIERS.get(row["frequency"], 1)
                        amt = avg / old_mult if old_mult else avg
        mult = FREQ_MULTIPLIERS.get(freq or "monthly", 1)
        amounts = {m: round((amt or 0) * mult, 2) for m in ML}
        fields.append("amounts=?"); vals.append(json.dumps(amounts))
    if not fields: return {"ok":True}
    vals.append(xid)
    with get_db() as db:
        db.execute(f"UPDATE expenses SET {','.join(fields)} WHERE id=?", vals)
    return {"ok":True}

@app.delete("/api/expenses/{xid}")
def del_expense(xid:str, user=Depends(require_admin)):
    with get_db() as db: db.execute("DELETE FROM expenses WHERE id=?", (xid,))
    return {"ok":True}

@app.get("/api/audit/{entity_id}")
def get_audit(entity_id:str, user=Depends(get_current_user)):
    with get_db() as db:
        rows = db.execute("SELECT * FROM audit_log WHERE entity_id=? ORDER BY changed_at DESC LIMIT 50", (entity_id,)).fetchall()
    return [dict(r) for r in rows]

class RateUpdate(BaseModel):
    rate: float

@app.get("/api/exchange-rate")
def api_get_rate():
    rate = get_exchange_rate()
    with get_db() as db:
        rows = db.execute("SELECT month, rate, locked FROM monthly_rates ORDER BY month").fetchall()
    monthly = {r["month"]: {"rate": r["rate"], "locked": bool(r["locked"])} for r in rows}
    now = datetime.now()
    current_month = f"{now.month:02d}-{now.year}"
    return {"rate": rate, "currentMonth": current_month, "monthlyRates": monthly}

@app.put("/api/exchange-rate")
def api_set_rate(b: RateUpdate, user=Depends(require_admin)):
    set_exchange_rate(b.rate)
    # Also set for current and future unlocked months
    now = datetime.now()
    current_month = f"{now.month:02d}-{now.year}"
    with get_db() as db:
        for m in ML:
            if m >= current_month:
                db.execute("INSERT OR REPLACE INTO monthly_rates (month, rate, locked) VALUES (?,?,0)", (m, b.rate))
    return {"ok": True, "rate": b.rate}

class MonthRateLock(BaseModel):
    month: str; rate: float

@app.put("/api/exchange-rate/lock")
def lock_month_rate(b: MonthRateLock, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("INSERT OR REPLACE INTO monthly_rates (month, rate, locked) VALUES (?,?,1)", (b.month, b.rate))
    return {"ok": True}

# ══════════════════════════════════════════
# INVOICES (Vendor Payments)
# ══════════════════════════════════════════
@app.get("/api/invoices")
def get_invoices(user=Depends(require_admin)):
    with get_db() as db:
        rows = db.execute("""
            SELECT i.*, e.vendor, e.dept FROM invoices i
            JOIN expenses e ON i.expense_id = e.id
            ORDER BY i.due_date DESC
        """).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        result.append({
            "id": d["id"], "expenseId": d["expense_id"],
            "vendor": d["vendor"], "dept": d["dept"],
            "invoiceNumber": d["invoice_number"],
            "amount": d["amount"], "currency": d["currency"],
            "dueDate": d["due_date"], "paidDate": d["paid_date"],
            "paidAmount": d["paid_amount"], "status": d["status"],
            "notes": d["notes"], "createdAt": d["created_at"]
        })
    # Compute vendor totals (unpaid)
    vendor_totals = {}
    for inv in result:
        v = inv["vendor"]
        if v not in vendor_totals:
            vendor_totals[v] = 0
        if inv["status"] != 'paid':
            vendor_totals[v] += to_usd(inv["amount"], inv["currency"])
    for inv in result:
        inv["vendorUnpaidTotal"] = round(vendor_totals.get(inv["vendor"], 0), 2)
        remaining = inv["amount"] - inv["paidAmount"]
        inv["remaining"] = round(remaining, 2)
        inv["remainingUsd"] = round(to_usd(remaining, inv["currency"]), 2)
    return result

class InvoiceIn(BaseModel):
    expenseId: str
    invoiceNumber: str = ""
    amount: float
    currency: str = "ILS"
    dueDate: Optional[str] = None
    notes: str = ""

@app.post("/api/invoices")
def add_invoice(b: InvoiceIn, user=Depends(require_admin)):
    iid = str(uuid4())[:8]
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("""INSERT INTO invoices (id, expense_id, invoice_number, amount, currency, due_date, status, notes, created_at)
                      VALUES (?,?,?,?,?,?,?,?,?)""",
                   (iid, b.expenseId, b.invoiceNumber, b.amount, b.currency, b.dueDate, "pending", b.notes, now))
    return {"ok": True, "id": iid}

class InvoiceUpdate(BaseModel):
    invoiceNumber: Optional[str] = None
    amount: Optional[float] = None
    currency: Optional[str] = None
    dueDate: Optional[str] = None
    paidDate: Optional[str] = None
    paidAmount: Optional[float] = None
    status: Optional[str] = None
    notes: Optional[str] = None

@app.put("/api/invoices/{iid}")
def upd_invoice(iid: str, b: InvoiceUpdate, user=Depends(require_admin)):
    fields = []; vals = []
    if b.invoiceNumber is not None: fields.append("invoice_number=?"); vals.append(b.invoiceNumber)
    if b.amount is not None: fields.append("amount=?"); vals.append(b.amount)
    if b.currency is not None: fields.append("currency=?"); vals.append(b.currency)
    if b.dueDate is not None: fields.append("due_date=?"); vals.append(b.dueDate)
    if b.paidDate is not None: fields.append("paid_date=?"); vals.append(b.paidDate if b.paidDate else None)
    if b.paidAmount is not None: fields.append("paid_amount=?"); vals.append(b.paidAmount)
    if b.status is not None: fields.append("status=?"); vals.append(b.status)
    if b.notes is not None: fields.append("notes=?"); vals.append(b.notes)
    # Auto-set status based on payment
    if b.paidAmount is not None and b.paidAmount > 0:
        with get_db() as db:
            row = db.execute("SELECT amount FROM invoices WHERE id=?", (iid,)).fetchone()
            if row:
                if b.paidAmount >= row["amount"]:
                    fields.append("status=?"); vals.append("paid")
                else:
                    fields.append("status=?"); vals.append("partial")
    if not fields: return {"ok": True}
    vals.append(iid)
    with get_db() as db:
        db.execute(f"UPDATE invoices SET {','.join(fields)} WHERE id=?", vals)
    return {"ok": True}

@app.delete("/api/invoices/{iid}")
def del_invoice(iid: str, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("DELETE FROM invoices WHERE id=?", (iid,))
    return {"ok": True}

# ══════════════════════════════════════════
# SALES TARGETS & KPIs
# ══════════════════════════════════════════
@app.get("/api/sales-targets")
def get_sales_targets(year: int = 2026, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "sales_manager"):
        raise HTTPException(403)
    with get_db() as db:
        rows = db.execute("SELECT * FROM sales_targets WHERE year=? ORDER BY salesperson_id, month", (year,)).fetchall()
    return [{"id":r["id"],"salespersonId":r["salesperson_id"],"year":r["year"],"month":r["month"],
             "revenueTarget":r["revenue_target"],"dealsTarget":r["deals_target"]} for r in rows]

class TargetSet(BaseModel):
    salespersonId: str
    year: int = 2026
    month: int  # 1-12
    revenueTarget: float = 0
    dealsTarget: int = 0
    avgDealTarget: float = 0
    cycleDaysTarget: int = 0
    closeRateTarget: float = 0
    newClientsTarget: int = 0

@app.post("/api/sales-targets")
def set_sales_target(b: TargetSet, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "sales_manager"):
        raise HTTPException(403)
    tid = f"st_{uuid4().hex[:8]}"
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("DELETE FROM sales_targets WHERE salesperson_id=? AND year=? AND month=?",
                   (b.salespersonId, b.year, b.month))
        db.execute("INSERT INTO sales_targets (id,salesperson_id,year,month,revenue_target,deals_target,avg_deal_target,cycle_days_target,close_rate_target,new_clients_target,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                   (tid, b.salespersonId, b.year, b.month, b.revenueTarget, b.dealsTarget, b.avgDealTarget, b.cycleDaysTarget, b.closeRateTarget, b.newClientsTarget, now))
    return {"ok": True}

@app.post("/api/sales-targets/bulk")
def set_bulk_targets(targets: list[TargetSet], user=Depends(get_current_user)):
    if user["role"] not in ("admin", "sales_manager"):
        raise HTTPException(403)
    now = datetime.now().isoformat()
    with get_db() as db:
        for b in targets:
            db.execute("DELETE FROM sales_targets WHERE salesperson_id=? AND year=? AND month=?",
                       (b.salespersonId, b.year, b.month))
            tid = f"st_{uuid4().hex[:8]}"
            db.execute("INSERT INTO sales_targets (id,salesperson_id,year,month,revenue_target,deals_target,avg_deal_target,cycle_days_target,close_rate_target,new_clients_target,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                       (tid, b.salespersonId, b.year, b.month, b.revenueTarget, b.dealsTarget, b.avgDealTarget, b.cycleDaysTarget, b.closeRateTarget, b.newClientsTarget, now))
    return {"ok": True}

@app.get("/api/sales-kpis")
def sales_kpis(year: int = 2026, user=Depends(get_current_user)):
    try:
        contracts = load_contracts()
        salespeople = load_salespeople()
    except Exception as e:
        return {"error": str(e)}
    with get_db() as db:
        targets = db.execute("SELECT * FROM sales_targets WHERE year=? ORDER BY salesperson_id, month", (year,)).fetchall()
    
    # Build targets lookup
    tgt_map = {}
    for t in targets:
        tgt_map.setdefault(t["salesperson_id"], {})[t["month"]] = {
            "rev": t["revenue_target"], "deals": t["deals_target"],
            "avgDeal": t["avg_deal_target"], "cycleDays": t["cycle_days_target"],
            "closeRate": t["close_rate_target"], "newClients": t["new_clients_target"]
        }
    
    # Stage order for conversion tracking
    stage_order = ["Initial Contact", "Email Sent", "Demo", "Negotiation", "Signed"]
    
    result = []
    for sp in salespeople:
        sid = sp["id"]
        sp_contracts = [c for c in contracts if c["salesperson_id"] == sid]
        sp_targets = tgt_map.get(sid, {})
        
        months_data = []
        for m in range(1, 13):
            m_contracts = [c for c in sp_contracts if c.get("sm") == m]
            signed = [c for c in m_contracts if c["stage"] == "Signed"]
            signed_rev = sum(to_usd(c["val"], c.get("currency", "USD")) for c in signed)
            signed_count = len(signed)
            total_count = len(m_contracts)
            
            # New clients
            new_clients = [c for c in signed if c.get("is_new_client", 1)]
            new_client_count = len(new_clients)
            new_client_rev = sum(to_usd(c["val"], c.get("currency", "USD")) for c in new_clients)
            
            # Avg deal size
            avg_deal = (signed_rev / signed_count) if signed_count > 0 else 0
            
            # Cycle time (days from created_at to stage_updated_at for signed deals)
            cycle_days_list = []
            for c in signed:
                if c.get("created_at") and c.get("stage_updated_at"):
                    try:
                        created = datetime.fromisoformat(c["created_at"].replace("Z",""))
                        closed = datetime.fromisoformat(c["stage_updated_at"].replace("Z",""))
                        cycle_days_list.append((closed - created).days)
                    except: pass
            avg_cycle = (sum(cycle_days_list) / len(cycle_days_list)) if cycle_days_list else 0
            
            # Close rate
            close_rate = (signed_count / total_count * 100) if total_count > 0 else 0
            
            # Stage-to-stage conversion
            stage_counts = {}
            for st in stage_order:
                st_idx = stage_order.index(st)
                stage_counts[st] = len([c for c in m_contracts if c["stage"] in stage_order and stage_order.index(c["stage"]) >= st_idx])
            
            conversions = {}
            for i in range(len(stage_order)-1):
                fr = stage_order[i]; to = stage_order[i+1]
                fr_count = stage_counts.get(fr, 0)
                to_count = stage_counts.get(to, 0)
                conversions[fr+"→"+to] = round(to_count / fr_count * 100, 1) if fr_count > 0 else 0
            
            # Targets
            tgt = sp_targets.get(m, {"rev":0,"deals":0,"avgDeal":0,"cycleDays":0,"closeRate":0,"newClients":0})
            
            months_data.append({
                "month": m,
                "signedRevenue": round(signed_rev, 2),
                "signedDeals": signed_count,
                "totalDeals": total_count,
                "avgDealSize": round(avg_deal, 2),
                "avgCycleDays": round(avg_cycle, 1),
                "closeRate": round(close_rate, 1),
                "newClients": new_client_count,
                "newClientRevenue": round(new_client_rev, 2),
                "newClientPct": round(new_client_rev / signed_rev * 100, 1) if signed_rev > 0 else 0,
                "conversions": conversions,
                "revenueTarget": tgt["rev"],
                "dealsTarget": tgt["deals"],
                "avgDealTarget": tgt["avgDeal"],
                "cycleDaysTarget": tgt["cycleDays"],
                "closeRateTarget": tgt["closeRate"],
                "newClientsTarget": tgt["newClients"],
                "revenuePct": round(signed_rev / tgt["rev"] * 100, 1) if tgt["rev"] > 0 else 0,
                "dealsPct": round(signed_count / tgt["deals"] * 100, 1) if tgt["deals"] > 0 else 0,
            })
        
        # Totals
        t_rev = sum(md["signedRevenue"] for md in months_data)
        t_deals = sum(md["signedDeals"] for md in months_data)
        t_total = sum(md["totalDeals"] for md in months_data)
        t_new = sum(md["newClients"] for md in months_data)
        t_new_rev = sum(md["newClientRevenue"] for md in months_data)
        t_rev_tgt = sum(md["revenueTarget"] for md in months_data)
        t_deals_tgt = sum(md["dealsTarget"] for md in months_data)
        t_new_tgt = sum(md["newClientsTarget"] for md in months_data)
        cycles = [md["avgCycleDays"] for md in months_data if md["avgCycleDays"] > 0]
        
        result.append({
            "salespersonId": sid,
            "salespersonName": sp["display_name"],
            "months": months_data,
            "totals": {
                "signedRevenue": round(t_rev, 2),
                "signedDeals": t_deals,
                "totalDeals": t_total,
                "avgDealSize": round(t_rev / t_deals, 2) if t_deals > 0 else 0,
                "avgCycleDays": round(sum(cycles) / len(cycles), 1) if cycles else 0,
                "closeRate": round(t_deals / t_total * 100, 1) if t_total > 0 else 0,
                "newClients": t_new,
                "newClientRevenue": round(t_new_rev, 2),
                "newClientPct": round(t_new_rev / t_rev * 100, 1) if t_rev > 0 else 0,
                "revenueTarget": t_rev_tgt,
                "dealsTarget": t_deals_tgt,
                "newClientsTarget": t_new_tgt,
                "revenuePct": round(t_rev / t_rev_tgt * 100, 1) if t_rev_tgt > 0 else 0,
                "dealsPct": round(t_deals / t_deals_tgt * 100, 1) if t_deals_tgt > 0 else 0,
            }
        })
    
    return result

# ══════════════════════════════════════════
# ══════════════════════════════════════════
# ALERTS SYSTEM
# ══════════════════════════════════════════
@app.get("/api/alerts")
def get_alerts(user=Depends(get_current_user)):
    uid = user["sub"]
    role = user["role"]
    now = datetime.now()
    alerts = []
    
    with get_db() as db:
        # Get dismissed alerts for this user
        dismissed = set()
        for r in db.execute("SELECT alert_key FROM alert_dismissals WHERE dismissed_by=?", (uid,)).fetchall():
            dismissed.add(r["alert_key"])
        
        # --- ALERT TYPE 1: Deal cycle time exceeded ---
        # Deals not yet Signed that have been open too long
        cycle_thresholds = {"Government": 90, "Enterprise": 60, "Academy": 45}
        contracts = db.execute("""SELECT c.id, c.client, c.country, c.industry, c.stage, c.created_at, 
                                        c.salesperson_id, u.display_name as sp_name
                                 FROM contracts c LEFT JOIN users u ON c.salesperson_id=u.id
                                 WHERE c.stage != 'Signed'""").fetchall()
        for c in contracts:
            if not c["created_at"]: continue
            try:
                created = datetime.fromisoformat(c["created_at"].replace("Z",""))
                days = (now - created).days
                threshold = cycle_thresholds.get(c["industry"], 60)
                if days > threshold:
                    key = f"cycle_{c['id']}_{days//7}"  # refreshes weekly so dismiss lasts ~1 week
                    if key not in dismissed:
                        # Filter by role: sales see only their own
                        if role == "sales" and c["salesperson_id"] != uid:
                            continue
                        alerts.append({
                            "key": key,
                            "type": "cycle_overdue",
                            "severity": "warning" if days < threshold * 1.5 else "critical",
                            "title": f"Deal overdue: {c['client']}",
                            "message": f"{days} days open (limit: {threshold}d for {c['industry']}). Stage: {c['stage']}",
                            "entityType": "contract",
                            "entityId": c["id"],
                            "assignee": c["sp_name"] or "Unassigned",
                            "assigneeId": c["salesperson_id"],
                            "daysOverdue": days - threshold,
                            "action": "Contact client and update status"
                        })
            except: pass
        
        # --- ALERT TYPE 2: Overdue invoices ---
        if role in ("admin", "sales_manager"):
            invoices = db.execute("""SELECT i.id, i.expense_id, i.invoice_number, i.amount, i.currency,
                                           i.due_date, i.paid_amount, i.status, e.vendor
                                    FROM invoices i LEFT JOIN expenses e ON i.expense_id=e.id
                                    WHERE i.status != 'paid' AND i.due_date IS NOT NULL""").fetchall()
            for inv in invoices:
                try:
                    due = datetime.fromisoformat(inv["due_date"])
                    if now > due:
                        days_late = (now - due).days
                        remaining = inv["amount"] - (inv["paid_amount"] or 0)
                        key = f"invoice_{inv['id']}_{days_late//7}"
                        if key not in dismissed:
                            cur_sym = "₪" if inv["currency"] == "ILS" else "$"
                            alerts.append({
                                "key": key,
                                "type": "invoice_overdue",
                                "severity": "warning" if days_late < 30 else "critical",
                                "title": f"Overdue payment: {inv['vendor']}",
                                "message": f"Invoice {inv['invoice_number'] or '#'} — {cur_sym}{remaining:,.0f} remaining. {days_late} days past due.",
                                "entityType": "invoice",
                                "entityId": inv["id"],
                                "assignee": "Finance",
                                "daysOverdue": days_late,
                                "action": "Update payment date or record payment"
                            })
                except: pass
        
        # --- ALERT TYPE 3: Stale deals (no update in 14+ days) ---
        stale_days = 14
        stale_contracts = db.execute("""SELECT c.id, c.client, c.stage, c.updated_at, c.salesperson_id,
                                              u.display_name as sp_name
                                       FROM contracts c LEFT JOIN users u ON c.salesperson_id=u.id
                                       WHERE c.stage != 'Signed'""").fetchall()
        for c in stale_contracts:
            if not c["updated_at"]: continue
            try:
                updated = datetime.fromisoformat(c["updated_at"].replace("Z",""))
                days_stale = (now - updated).days
                if days_stale >= stale_days:
                    key = f"stale_{c['id']}_{days_stale//7}"
                    if key not in dismissed:
                        if role == "sales" and c["salesperson_id"] != uid:
                            continue
                        alerts.append({
                            "key": key,
                            "type": "stale_deal",
                            "severity": "info" if days_stale < 30 else "warning",
                            "title": f"Stale deal: {c['client']}",
                            "message": f"No updates for {days_stale} days. Stage: {c['stage']}",
                            "entityType": "contract",
                            "entityId": c["id"],
                            "assignee": c["sp_name"] or "Unassigned",
                            "assigneeId": c.get("salesperson_id"),
                            "daysOverdue": days_stale - stale_days,
                            "action": "Follow up with client"
                        })
            except: pass
    
    # Sort by severity (critical first), then by daysOverdue
    sev_order = {"critical": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda a: (sev_order.get(a["severity"], 9), -a.get("daysOverdue", 0)))
    
    return {"alerts": alerts, "count": len(alerts)}

class DismissAlert(BaseModel):
    alertKey: str
    actionNote: str = ""

@app.post("/api/alerts/dismiss")
def dismiss_alert(b: DismissAlert, user=Depends(get_current_user)):
    uid = user["sub"]
    now = datetime.now().isoformat()
    aid = f"ad_{uuid4().hex[:8]}"
    with get_db() as db:
        db.execute("INSERT OR REPLACE INTO alert_dismissals (id, alert_key, dismissed_by, dismissed_at, action_note) VALUES (?,?,?,?,?)",
                   (aid, b.alertKey, uid, now, b.actionNote))
    return {"ok": True}

class TaxPayment(BaseModel):
    month: int  # 1-12
    amount: float

@app.post("/api/tax-payment")
def set_tax_payment(b: TaxPayment, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES (?, ?)",
                   (f"tax_{b.month}", str(b.amount)))
    return {"ok": True}

# ══════════════════════════════════════════
# LOANS
# ══════════════════════════════════════════
def load_loans():
    with get_db() as db:
        rows = db.execute("SELECT * FROM loans ORDER BY start_date").fetchall()
    return [dict(r) for r in rows]

def calc_loan_schedule(amount, term_months, annual_rate, start_date_str, repayment_type, interest_start_month=""):
    monthly_rate = annual_rate / 100 / 12
    schedule = []
    remaining = amount
    try: sd = date.fromisoformat(start_date_str)
    except: sd = date(2026, 1, 1)
    
    # Determine when interest starts
    interest_start = None
    if interest_start_month:
        try:
            parts = interest_start_month.split("-")
            interest_start = date(int(parts[1]), int(parts[0]), 1)
        except: pass
    
    for i in range(term_months):
        m = (sd.month + i - 1) % 12 + 1
        y = sd.year + (sd.month + i - 1) // 12
        pay_date = date(y, m, min(sd.day, 28))
        
        # Check if interest has started yet
        charge_interest = True
        if interest_start and pay_date < interest_start:
            charge_interest = False
        
        interest = remaining * monthly_rate if charge_interest else 0
        if repayment_type == 'interest_only':
            principal = remaining if i == term_months - 1 else 0
            payment = interest + principal
        elif repayment_type == 'principal_only':
            principal = amount / term_months; interest = 0; payment = principal
        elif repayment_type == 'balloon':
            # Grace period: pay nothing until last month, then pay everything
            principal = remaining if i == term_months - 1 else 0
            payment = principal + (interest if i == term_months - 1 else 0)
            if i < term_months - 1:
                interest = 0; payment = 0  # true grace — no payments at all
        else:
            if monthly_rate > 0:
                payment = amount * monthly_rate / (1 - (1 + monthly_rate) ** (-term_months))
            else:
                payment = amount / term_months
            principal = payment - interest
        remaining = max(0, remaining - principal)
        schedule.append({"month": i+1, "date": pay_date.isoformat(), "payment": round(payment,2),
                         "principal": round(principal,2), "interest": round(interest,2), "remaining": round(remaining,2)})
    return schedule

@app.get("/api/loans")
def get_loans(user=Depends(require_admin)):
    loans = load_loans()
    result = []
    for ln in loans:
        sch = calc_loan_schedule(ln["amount"], ln["term_months"], ln["annual_rate"], ln["start_date"], ln["repayment_type"], ln.get("interest_start_month",""))
        result.append({"id":ln["id"],"lender":ln["lender"],"amount":ln["amount"],
            "termMonths":ln["term_months"],"annualRate":ln["annual_rate"],
            "startDate":ln["start_date"],"repaymentType":ln["repayment_type"],
            "currency":ln.get("currency","ILS"),"notes":ln.get("notes",""),
            "totalPayment":round(sum(s["payment"] for s in sch),2),
            "totalInterest":round(sum(s["interest"] for s in sch),2),
            "schedule":sch})
    return result

class LoanIn(BaseModel):
    lender:str; amount:float; termMonths:int; annualRate:float=0
    startDate:str; repaymentType:str="principal_interest"
    currency:str="ILS"; notes:str=""
    interestStartMonth:str=""

@app.post("/api/loans")
def add_loan(b: LoanIn, user=Depends(require_admin)):
    lid = f"ln_{uuid4().hex[:8]}"
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("INSERT INTO loans (id,lender,amount,term_months,annual_rate,start_date,repayment_type,currency,notes,interest_start_month,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                   (lid,b.lender,b.amount,b.termMonths,b.annualRate,b.startDate,b.repaymentType,b.currency,b.notes,b.interestStartMonth,now))
    return {"ok":True,"id":lid}

@app.put("/api/loans/{lid}")
def upd_loan(lid:str, b:LoanIn, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("UPDATE loans SET lender=?,amount=?,term_months=?,annual_rate=?,start_date=?,repayment_type=?,currency=?,notes=?,interest_start_month=? WHERE id=?",
                   (b.lender,b.amount,b.termMonths,b.annualRate,b.startDate,b.repaymentType,b.currency,b.notes,b.interestStartMonth,lid))
    return {"ok":True}

@app.delete("/api/loans/{lid}")
def del_loan(lid:str, user=Depends(require_admin)):
    with get_db() as db: db.execute("DELETE FROM loans WHERE id=?",(lid,))
    return {"ok":True}

class LoanOpeningBalance(BaseModel):
    amount: float

@app.put("/api/loans/opening-balance")
def set_loan_opening_balance(b: LoanOpeningBalance, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('loan_opening_balance', ?)", (str(b.amount),))
    return {"ok": True}

# ══════════════════════════════════════════
# CAP TABLE
# ══════════════════════════════════════════
@app.get("/api/cap-table")
def get_cap_table(user=Depends(require_admin)):
    with get_db() as db:
        rows = db.execute("SELECT * FROM cap_table ORDER BY shares DESC").fetchall()
    entries = [dict(r) for r in rows]
    total_shares = sum(e["shares"] for e in entries)
    esop_shares = sum(e["shares"] for e in entries if e["is_esop"])
    return {"entries": [{"id":e["id"],"holderName":e["holder_name"],"shares":e["shares"],
                         "isEsop":bool(e["is_esop"]),"notes":e.get("notes",""),
                         "pct":round(e["shares"]/total_shares*100,2) if total_shares>0 else 0} for e in entries],
            "totalShares":total_shares, "esopShares":esop_shares,
            "esopPct":round(esop_shares/total_shares*100,2) if total_shares>0 else 0}

class CapEntry(BaseModel):
    holderName:str; shares:float; isEsop:bool=False; notes:str=""

@app.post("/api/cap-table")
def add_cap_entry(b: CapEntry, user=Depends(require_admin)):
    cid = f"cap_{uuid4().hex[:8]}"
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("INSERT INTO cap_table (id,holder_name,shares,is_esop,notes,created_at) VALUES (?,?,?,?,?,?)",
                   (cid, b.holderName, b.shares, int(b.isEsop), b.notes, now))
    return {"ok":True,"id":cid}

@app.put("/api/cap-table/{cid}")
def upd_cap_entry(cid:str, b:CapEntry, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("UPDATE cap_table SET holder_name=?,shares=?,is_esop=?,notes=? WHERE id=?",
                   (b.holderName, b.shares, int(b.isEsop), b.notes, cid))
    return {"ok":True}

@app.delete("/api/cap-table/{cid}")
def del_cap_entry(cid:str, user=Depends(require_admin)):
    with get_db() as db: db.execute("DELETE FROM cap_table WHERE id=?",(cid,))
    return {"ok":True}

# ══════════════════════════════════════════
# USER PREFERENCES (layout order)
# ══════════════════════════════════════════
@app.get("/api/preferences")
def get_preferences(user=Depends(get_current_user)):
    with get_db() as db:
        rows = db.execute("SELECT pref_key, pref_value FROM user_preferences WHERE user_id=?", (user["sub"],)).fetchall()
    return {r["pref_key"]: json.loads(r["pref_value"]) for r in rows}

class PrefUpdate(BaseModel):
    key: str
    value: list

@app.put("/api/preferences")
def set_preference(b: PrefUpdate, user=Depends(get_current_user)):
    with get_db() as db:
        db.execute("INSERT OR REPLACE INTO user_preferences (user_id, pref_key, pref_value) VALUES (?,?,?)",
                   (user["sub"], b.key, json.dumps(b.value)))
    return {"ok": True}

# ══════════════════════════════════════════
# ACTION ITEMS
# ══════════════════════════════════════════
class ActionItemIn(BaseModel):
    contractId: str
    description: str
    dueDate: str

@app.post("/api/actions")
def add_action(b: ActionItemIn, user=Depends(get_current_user)):
    aid = f"act_{uuid4().hex[:8]}"
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("INSERT INTO action_items (id,contract_id,user_id,description,due_date,created_at) VALUES (?,?,?,?,?,?)",
                   (aid, b.contractId, user["sub"], b.description, b.dueDate, now))
    return {"ok": True, "id": aid}

@app.put("/api/actions/{aid}/complete")
def complete_action(aid: str, user=Depends(get_current_user)):
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("UPDATE action_items SET completed=1, completed_at=? WHERE id=? AND user_id=?", (now, aid, user["sub"]))
    return {"ok": True}

@app.delete("/api/actions/{aid}")
def delete_action(aid: str, user=Depends(get_current_user)):
    with get_db() as db:
        db.execute("DELETE FROM action_items WHERE id=? AND user_id=?", (aid, user["sub"]))
    return {"ok": True}

@app.get("/api/actions/today")
def get_today_actions(user=Depends(get_current_user)):
    today = date.today().isoformat()
    with get_db() as db:
        role = db.execute("SELECT role, can_see_all FROM users WHERE id=?", (user["sub"],)).fetchone()
        is_manager = role and role["role"] in ("admin", "sales_manager")
        
        if is_manager:
            # Manager sees everyone's actions
            rows = db.execute("""SELECT a.*, c.client, u.display_name as assignee FROM action_items a 
                                LEFT JOIN contracts c ON a.contract_id=c.id
                                LEFT JOIN users u ON a.user_id=u.id
                                WHERE a.completed=0 AND a.due_date<=?
                                ORDER BY a.due_date""", (today,)).fetchall()
        else:
            # Sales rep sees only their own
            rows = db.execute("""SELECT a.*, c.client, u.display_name as assignee FROM action_items a 
                                LEFT JOIN contracts c ON a.contract_id=c.id
                                LEFT JOIN users u ON a.user_id=u.id
                                WHERE a.user_id=? AND a.completed=0 AND a.due_date<=?
                                ORDER BY a.due_date""", (user["sub"], today)).fetchall()
    return [{"id":r["id"],"contractId":r["contract_id"],"client":r["client"] or "Unknown",
             "description":r["description"],"dueDate":r["due_date"],
             "assignee":r["assignee"] or "Unknown","userId":r["user_id"],
             "isOverdue":r["due_date"]<today,"isToday":r["due_date"]==today} for r in rows]

@app.get("/api/actions/contract/{cid}")
def get_contract_actions(cid: str, user=Depends(get_current_user)):
    with get_db() as db:
        rows = db.execute("""SELECT * FROM action_items WHERE contract_id=? AND user_id=? 
                            ORDER BY completed, due_date""", (cid, user["sub"])).fetchall()
    return [{"id":r["id"],"description":r["description"],"dueDate":r["due_date"],
             "completed":bool(r["completed"]),"completedAt":r["completed_at"]} for r in rows]

# ══════════════════════════════════════════
# ══════════════════════════════════════════
# EXPENSE PAYMENTS (per month payment dates)
# ══════════════════════════════════════════
class ExpensePayment(BaseModel):
    expenseId: str
    amount: float
    invoiceDate: str = ""
    paymentDate: str = ""
    notes: str = ""

@app.post("/api/expense-payment")
def save_expense_payment(b: ExpensePayment, user=Depends(require_admin)):
    """Save a single invoice entry for an expense. Multiple allowed per month."""
    epid = f"ep_{uuid4().hex[:8]}"
    now = datetime.now().isoformat()
    # Determine month from invoice_date
    month = ""
    if b.invoiceDate:
        try:
            d2 = date.fromisoformat(b.invoiceDate)
            month = f"{d2.month:02d}-{d2.year}"
        except: pass
    with get_db() as db:
        db.execute("""INSERT INTO expense_payments 
                     (id, expense_id, month, amount, invoice_date, payment_date, notes, created_at)
                     VALUES (?,?,?,?,?,?,?,?)""",
                   (epid, b.expenseId, month, b.amount, b.invoiceDate, b.paymentDate, b.notes, now))
    return {"ok": True, "id": epid}

@app.delete("/api/expense-payment/{epid}")
def delete_expense_payment(epid: str, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("DELETE FROM expense_payments WHERE id=?", (epid,))
    return {"ok": True}

@app.get("/api/expense-payments/{expense_id}")
def get_expense_payments(expense_id: str, user=Depends(require_admin)):
    with get_db() as db:
        rows = db.execute("SELECT * FROM expense_payments WHERE expense_id=? ORDER BY invoice_date", (expense_id,)).fetchall()
    return [{"id":r["id"],"month":r["month"],"amount":r["amount"],
             "invoiceDate":r["invoice_date"] or "","paymentDate":r["payment_date"] or "",
             "notes":r["notes"] or ""} for r in rows]

# SALARY SNAPSHOTS
# ══════════════════════════════════════════

# ══════════════════════════════════════════
# INCOME SOURCES & PAYMENTS
# ══════════════════════════════════════════
class IncomeSourceIn(BaseModel):
    client: str
    category: str = ""
    currency: str = "ILS"
    contactName: str = ""
    contactPhone: str = ""
    contactEmail: str = ""
    isRecurring: bool = False
    subjectToVat: bool = False
    notes: str = ""

@app.post("/api/income")
def add_income_source(b: IncomeSourceIn, user=Depends(require_admin)):
    iid = f"inc_{uuid4().hex[:8]}"
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("""INSERT INTO income_sources (id,client,category,currency,contact_name,contact_phone,
                      contact_email,is_recurring,subject_to_vat,notes,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                   (iid, b.client, b.category, b.currency, b.contactName, b.contactPhone,
                    b.contactEmail, int(b.isRecurring), int(b.subjectToVat), b.notes, now))
    return {"ok": True, "id": iid}

class IncomeSourceUpdate(BaseModel):
    client: Optional[str] = None
    category: Optional[str] = None
    currency: Optional[str] = None
    contactName: Optional[str] = None
    contactPhone: Optional[str] = None
    contactEmail: Optional[str] = None
    isRecurring: Optional[bool] = None
    subjectToVat: Optional[bool] = None
    notes: Optional[str] = None

@app.put("/api/income/{iid}")
def upd_income_source(iid: str, b: IncomeSourceUpdate, user=Depends(require_admin)):
    fields = []; vals = []
    if b.client is not None: fields.append("client=?"); vals.append(b.client)
    if b.category is not None: fields.append("category=?"); vals.append(b.category)
    if b.currency is not None: fields.append("currency=?"); vals.append(b.currency)
    if b.contactName is not None: fields.append("contact_name=?"); vals.append(b.contactName)
    if b.contactPhone is not None: fields.append("contact_phone=?"); vals.append(b.contactPhone)
    if b.contactEmail is not None: fields.append("contact_email=?"); vals.append(b.contactEmail)
    if b.isRecurring is not None: fields.append("is_recurring=?"); vals.append(int(b.isRecurring))
    if b.subjectToVat is not None: fields.append("subject_to_vat=?"); vals.append(int(b.subjectToVat))
    if b.notes is not None: fields.append("notes=?"); vals.append(b.notes)
    if not fields: return {"ok": True}
    vals.append(iid)
    with get_db() as db:
        db.execute(f"UPDATE income_sources SET {','.join(fields)} WHERE id=?", vals)
    return {"ok": True}

@app.delete("/api/income/{iid}")
def del_income_source(iid: str, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("DELETE FROM income_sources WHERE id=?", (iid,))
        db.execute("DELETE FROM income_payments WHERE income_id=?", (iid,))
    return {"ok": True}

class IncomePaymentIn(BaseModel):
    incomeId: str
    amount: float
    invoiceDate: str = ""
    paymentDate: str = ""
    notes: str = ""

@app.post("/api/income-payment")
def add_income_payment(b: IncomePaymentIn, user=Depends(require_admin)):
    ipid = f"ip_{uuid4().hex[:8]}"
    now = datetime.now().isoformat()
    month = ""
    if b.invoiceDate:
        try:
            d2 = date.fromisoformat(b.invoiceDate)
            month = f"{d2.month:02d}-{d2.year}"
        except: pass
    with get_db() as db:
        db.execute("""INSERT INTO income_payments (id,income_id,amount,invoice_date,payment_date,notes,created_at)
                     VALUES (?,?,?,?,?,?,?)""",
                   (ipid, b.incomeId, b.amount, b.invoiceDate, b.paymentDate, b.notes, now))
    return {"ok": True, "id": ipid}

@app.delete("/api/income-payment/{ipid}")
def del_income_payment(ipid: str, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("DELETE FROM income_payments WHERE id=?", (ipid,))
    return {"ok": True}

@app.get("/api/income-payments/{income_id}")
def get_income_payments(income_id: str, user=Depends(require_admin)):
    with get_db() as db:
        rows = db.execute("SELECT * FROM income_payments WHERE income_id=? ORDER BY invoice_date", (income_id,)).fetchall()
    return [{"id":r["id"],"amount":r["amount"],
             "invoiceDate":r["invoice_date"] or "","paymentDate":r["payment_date"] or "",
             "notes":r["notes"] or ""} for r in rows]

# ══════════════════════════════════════════
class SnapshotMonth(BaseModel):
    month: str
    netPayDate: str = ""
    taxPayDate: str = ""
    pensionPayDate: str = ""

@app.post("/api/salary-snapshot")
def save_salary_snapshot(b: SnapshotMonth, user=Depends(require_admin)):
    """Snapshot current employee data for a specific month with payment dates."""
    employees = load_employees()
    now = datetime.now().isoformat()
    count = 0
    with get_db() as db:
        for e in employees:
            if e["term_date"]:
                try:
                    td = date.fromisoformat(e["term_date"])
                    p = b.month.split("-")
                    md = date(int(p[1]), int(p[0]), 1)
                    if md > td: continue
                except: pass
            try:
                hd = date.fromisoformat(e["hire_date"])
                p = b.month.split("-")
                md = date(int(p[1]), int(p[0]), 1)
                if md < hd: continue
            except: pass
            
            sid = f"ss_{e['id']}_{b.month.replace('-','')}"
            db.execute("""INSERT OR REPLACE INTO salary_snapshots 
                         (id, employee_id, month, gross, position_pct, bonus, currency, dept,
                          pension_pct, disability_pct, study_fund_pct, study_fund_salary,
                          net_pay_date, tax_pay_date, pension_pay_date, created_at)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                       (sid, e["id"], b.month, e["gross"], e.get("position_pct", 100),
                        e.get("bonus", 0), e.get("currency", "ILS"), e["dept"],
                        e.get("pension_pct", 8.33), e.get("disability_pct", 2.5),
                        e.get("study_fund_pct", 7.5), e.get("study_fund_salary", 0),
                        b.netPayDate, b.taxPayDate, b.pensionPayDate, now))
            count += 1
    return {"ok": True, "count": count, "month": b.month}

@app.get("/api/salary-snapshots")
def get_salary_snapshots(user=Depends(require_admin)):
    """Get list of months that have snapshots."""
    with get_db() as db:
        rows = db.execute("SELECT DISTINCT month, COUNT(*) as cnt, MAX(created_at) as saved_at FROM salary_snapshots GROUP BY month ORDER BY month").fetchall()
    return [{"month": r["month"], "count": r["cnt"], "savedAt": r["saved_at"]} for r in rows]

def load_employee_for_month(employee_id, month_str):
    """Get employee data for a specific month: snapshot if available, else current."""
    with get_db() as db:
        snap = db.execute("SELECT * FROM salary_snapshots WHERE employee_id=? AND month=?", 
                         (employee_id, month_str)).fetchone()
    if snap:
        return dict(snap)
    return None  # Caller uses current employee data

# SENSITIVITY ANALYSIS
# ══════════════════════════════════════════
class SensitivityScenario(BaseModel):
    name: str = "Scenario"
    revenueChangePct: float = 0
    expenseChangePct: float = 0
    fxChangePct: float = 0
    revenueDelayMonths: int = 0
    expenseDelayMonths: int = 0

@app.post("/api/sensitivity")
def compute_sensitivity(scenarios: list[SensitivityScenario], user=Depends(require_admin)):
    base = compute(0, 0, 15000, False)
    months_labels = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    base_data = {
        "labels": months_labels,
        "revenue": [round(m["revenue"], 2) for m in base],
        "cashIn": [round(m["cashIn"], 2) for m in base],
        "expenses": [round(m["totalExp"], 2) for m in base],
        "netIncome": [round(m["netIncome"], 2) for m in base],
        "cumCash": [round(m["cumCash"], 2) for m in base],
    }
    results = []
    for sc in scenarios:
        adj_revenue = list(base_data["revenue"])
        adj_cashIn = list(base_data["cashIn"])
        adj_expenses = list(base_data["expenses"])
        if sc.revenueChangePct != 0:
            f = 1 + sc.revenueChangePct / 100
            adj_revenue = [round(r * f, 2) for r in adj_revenue]
            adj_cashIn = [round(r * f, 2) for r in adj_cashIn]
        if sc.revenueDelayMonths > 0:
            d = sc.revenueDelayMonths
            adj_revenue = [0.0]*d + adj_revenue[:12-d]
            adj_cashIn = [0.0]*d + adj_cashIn[:12-d]
        if sc.expenseChangePct != 0:
            f = 1 + sc.expenseChangePct / 100
            adj_expenses = [round(e * f, 2) for e in adj_expenses]
        if sc.expenseDelayMonths > 0:
            d = sc.expenseDelayMonths
            adj_expenses = [0.0]*d + adj_expenses[:12-d]
        if sc.fxChangePct != 0:
            fx = 1 + sc.fxChangePct / 100
            adj_revenue = [round(r * (0.7 + 0.3 * fx), 2) for r in adj_revenue]
            adj_cashIn = [round(r * (0.7 + 0.3 * fx), 2) for r in adj_cashIn]
            adj_expenses = [round(e * (0.4 + 0.6 * fx), 2) for e in adj_expenses]
        adj_net = [round(adj_revenue[i] - adj_expenses[i], 2) for i in range(12)]
        cum = 0; adj_cum = []
        for i in range(12):
            cum += adj_cashIn[i] - adj_expenses[i]
            adj_cum.append(round(cum, 2))
        results.append({
            "name": sc.name, "revenue": adj_revenue, "cashIn": adj_cashIn,
            "expenses": adj_expenses, "netIncome": adj_net, "cumCash": adj_cum,
            "totals": {"revenue": round(sum(adj_revenue), 2), "expenses": round(sum(adj_expenses), 2),
                       "netIncome": round(sum(adj_net), 2), "endCash": adj_cum[-1] if adj_cum else 0}
        })
    return {"base": base_data, "scenarios": results, "baseTotals": {
        "revenue": round(sum(base_data["revenue"]), 2), "expenses": round(sum(base_data["expenses"]), 2),
        "netIncome": round(sum(base_data["netIncome"]), 2), "endCash": base_data["cumCash"][-1] if base_data["cumCash"] else 0}}

# SCENARIOS (What-If)
# ══════════════════════════════════════════
@app.get("/api/scenarios")
def list_scenarios(user=Depends(require_admin)):
    with get_db() as db:
        rows = db.execute("SELECT * FROM scenarios ORDER BY updated_at DESC").fetchall()
    return [{"id":r["id"],"name":r["name"],"description":r["description"],"updatedAt":r["updated_at"]} for r in rows]

class ScenarioCreate(BaseModel):
    name: str
    description: str = ""

@app.post("/api/scenarios")
def create_scenario(b: ScenarioCreate, user=Depends(require_admin)):
    sid = f"sc_{uuid4().hex[:8]}"
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("INSERT INTO scenarios (id,name,description,created_at,updated_at) VALUES (?,?,?,?,?)",
                   (sid, b.name, b.description, now, now))
    return {"ok": True, "id": sid}

@app.delete("/api/scenarios/{sid}")
def delete_scenario(sid: str, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("DELETE FROM scenario_items WHERE scenario_id=?", (sid,))
        db.execute("DELETE FROM scenarios WHERE id=?", (sid,))
    return {"ok": True}

@app.get("/api/scenarios/{sid}")
def get_scenario(sid: str, user=Depends(require_admin)):
    with get_db() as db:
        sc = db.execute("SELECT * FROM scenarios WHERE id=?", (sid,)).fetchone()
        if not sc: raise HTTPException(404)
        items = db.execute("SELECT * FROM scenario_items WHERE scenario_id=? ORDER BY id", (sid,)).fetchall()
    return {
        "id": sc["id"], "name": sc["name"], "description": sc["description"],
        "items": [{"id":it["id"],"type":it["item_type"],"config":json.loads(it["config"])} for it in items]
    }

class ScenarioItemCreate(BaseModel):
    itemType: str  # "hire", "expense_change", "revenue_stream", "ttm_reduction"
    config: dict

@app.post("/api/scenarios/{sid}/items")
def add_scenario_item(sid: str, b: ScenarioItemCreate, user=Depends(require_admin)):
    iid = f"si_{uuid4().hex[:8]}"
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("INSERT INTO scenario_items (id,scenario_id,item_type,config) VALUES (?,?,?,?)",
                   (iid, sid, b.itemType, json.dumps(b.config)))
        db.execute("UPDATE scenarios SET updated_at=? WHERE id=?", (now, sid))
    return {"ok": True, "id": iid}

@app.delete("/api/scenarios/{sid}/items/{iid}")
def remove_scenario_item(sid: str, iid: str, user=Depends(require_admin)):
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("DELETE FROM scenario_items WHERE id=? AND scenario_id=?", (iid, sid))
        db.execute("UPDATE scenarios SET updated_at=? WHERE id=?", (now, sid))
    return {"ok": True}

@app.put("/api/scenarios/{sid}/items/{iid}")
def update_scenario_item(sid: str, iid: str, b: ScenarioItemCreate, user=Depends(require_admin)):
    now = datetime.now().isoformat()
    with get_db() as db:
        db.execute("UPDATE scenario_items SET item_type=?, config=? WHERE id=? AND scenario_id=?",
                   (b.itemType, json.dumps(b.config), iid, sid))
        db.execute("UPDATE scenarios SET updated_at=? WHERE id=?", (now, sid))
    return {"ok": True}

@app.get("/api/scenarios/{sid}/compute")
def compute_scenario(sid: str, user=Depends(require_admin)):
    """Compute baseline vs scenario financials."""
    # Baseline
    base_summary = compute(0, 0, 15000, True)
    rate = get_exchange_rate()
    
    # Load scenario items
    with get_db() as db:
        items = db.execute("SELECT * FROM scenario_items WHERE scenario_id=?", (sid,)).fetchall()
    
    # Apply scenario adjustments on top of baseline
    scenario_months = []
    for i, bs in enumerate(base_summary):
        sm = dict(bs)
        sm["lab"] = dict(bs["lab"])
        sm["vd"] = dict(bs["vd"])
        sm["addedRev"] = 0
        sm["addedExp"] = 0
        scenario_months.append(sm)
    
    for it in items:
        cfg = json.loads(it["config"])
        itype = it["item_type"]
        
        if itype == "hire":
            dept = cfg.get("dept", "S&M")
            cost_orig = cfg.get("monthlyCost", 0)
            cur = cfg.get("currency", "ILS")
            total_hire, _ = calc_employer_cost(cost_orig, 0, 100, 8.33, 2.5, 7.5, 0, cur)
            start_m = cfg.get("startMonth", 1) - 1
            
            # Add labor cost from start month
            for mi in range(start_m, 12):
                cost_usd_m = to_usd(total_hire, cur, ML[mi])
                scenario_months[mi]["lab"][dept] = scenario_months[mi]["lab"].get(dept, 0) + cost_usd_m
                scenario_months[mi]["totalExp"] = scenario_months[mi].get("totalExp", 0) + cost_usd
                scenario_months[mi]["addedExp"] += cost_usd
            
            if dept == "R&D":
                # R&D hire: TTM reduction - pull future revenue forward
                ttm_months = cfg.get("ttmReduction", 0)
                if ttm_months > 0:
                    for mi in range(12):
                        future_mi = mi + ttm_months
                        if future_mi < 12:
                            pulled_rev = base_summary[future_mi]["cashIn"] - base_summary[mi]["cashIn"]
                            if pulled_rev > 0:
                                boost = pulled_rev * 0.5
                                scenario_months[mi]["cashIn"] = scenario_months[mi].get("cashIn", 0) + boost
                                scenario_months[mi]["addedRev"] += boost
            else:
                # S&M / other hire: direct revenue generation
                rev_per_month = cfg.get("expectedRevenue", 0)
                rev_start = cfg.get("revenueStartMonth", 7) - 1
                rev_growth = cfg.get("revenueGrowthPct", 0) / 100.0
                if rev_per_month > 0:
                    for mi in range(rev_start, 12):
                        months_active = mi - rev_start
                        growth_factor = 1 + (rev_growth * months_active / 12)
                        rev = rev_per_month * growth_factor
                        scenario_months[mi]["cashIn"] = scenario_months[mi].get("cashIn", 0) + rev
                        scenario_months[mi]["addedRev"] += rev
        
        elif itype == "expense_change":
            exp_cur = cfg.get("currency", "USD")
            dept = cfg.get("dept", "G&A")
            from_m = cfg.get("fromMonth", 1) - 1
            
            for mi in range(from_m, 12):
                new_amt = to_usd(cfg.get("newAmount", 0), exp_cur, ML[mi])
                old_amt = to_usd(cfg.get("oldAmount", 0), exp_cur, ML[mi])
                delta = new_amt - old_amt
                scenario_months[mi]["vd"][dept] = scenario_months[mi]["vd"].get(dept, 0) + delta
                scenario_months[mi]["totalExp"] = scenario_months[mi].get("totalExp", 0) + delta
                scenario_months[mi]["addedExp"] += delta
        
        elif itype == "revenue_stream":
            # New revenue: amount, fromMonth, growthPct
            amt = cfg.get("monthlyAmount", 0)
            from_m = cfg.get("fromMonth", 1) - 1
            growth = cfg.get("growthPct", 0) / 100.0
            
            for mi in range(from_m, 12):
                months_active = mi - from_m
                growth_factor = 1 + (growth * months_active / 12)
                rev = amt * growth_factor
                scenario_months[mi]["cashIn"] = scenario_months[mi].get("cashIn", 0) + rev
                scenario_months[mi]["addedRev"] += rev
    
    # Recalculate cumulative cash
    prev_cash = 0
    for sm in scenario_months:
        sm["totalExp"] = sum(sm["lab"].values()) + sum(sm["vd"].values())
        sm["net"] = sm["cashIn"] - sm["totalExp"]
        sm["cumCash"] = prev_cash + sm["net"]
        prev_cash = sm["cumCash"]
    
    # Build response
    labels = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    result = []
    base_prev = 0
    for i in range(12):
        bs = base_summary[i]
        sc = scenario_months[i]
        b_total_exp = sum(bs["lab"].values()) + sum(bs["vd"].values())
        b_net = bs["cashIn"] - b_total_exp
        base_cum = base_prev + b_net
        base_prev = base_cum
        
        result.append({
            "month": labels[i],
            "baseRevenue": round(bs["revenue"], 2),
            "baseExpenses": round(b_total_exp, 2),
            "baseNet": round(b_net, 2),
            "baseCash": round(base_cum, 2),
            "baseCashIn": round(bs["cashIn"], 2),
            "scenRevenue": round(sc.get("revenue", 0) + sc.get("addedRev", 0), 2),
            "scenExpenses": round(sc["totalExp"], 2),
            "scenNet": round(sc["net"], 2),
            "scenCash": round(sc["cumCash"], 2),
            "deltaRevenue": round((sc.get("revenue", 0) + sc.get("addedRev", 0)) - bs["revenue"], 2),
            "deltaExpenses": round(sc["totalExp"] - b_total_exp, 2),
            "deltaNet": round(sc["net"] - b_net, 2),
            "deltaCash": round(sc["cumCash"] - base_cum, 2),
            "addedRev": round(sc.get("addedRev", 0), 2),
            "addedExp": round(sc.get("addedExp", 0), 2),
        })
    
    # Totals
    totals = {
        "baseRevenue": round(sum(r["baseRevenue"] for r in result), 2),
        "baseExpenses": round(sum(r["baseExpenses"] for r in result), 2),
        "scenRevenue": round(sum(r["scenRevenue"] for r in result), 2),
        "scenExpenses": round(sum(r["scenExpenses"] for r in result), 2),
    }
    totals["deltaRevenue"] = round(totals["scenRevenue"] - totals["baseRevenue"], 2)
    totals["deltaExpenses"] = round(totals["scenExpenses"] - totals["baseExpenses"], 2)
    totals["baseNet"] = round(totals["baseRevenue"] - totals["baseExpenses"], 2)
    totals["scenNet"] = round(totals["scenRevenue"] - totals["scenExpenses"], 2)
    totals["deltaNet"] = round(totals["scenNet"] - totals["baseNet"], 2)
    
    return {"months": result, "totals": totals}

# ══════════════════════════════════════════
# PAYMENT METHODS
# ══════════════════════════════════════════
@app.get("/api/payment-methods")
def list_payment_methods(user=Depends(require_admin)):
    with get_db() as db:
        rows = db.execute("SELECT * FROM payment_methods ORDER BY sort_order").fetchall()
    return [{"id":r["id"],"name":r["name"],"bankGroup":r["bank_group"],"isActive":bool(r["is_active"])} for r in rows]

class PMCreate(BaseModel):
    name: str
    bankGroup: str = ""

@app.post("/api/payment-methods")
def create_payment_method(body: PMCreate, user=Depends(require_admin)):
    pid = f"pm_{uuid4().hex[:8]}"
    with get_db() as db:
        mx = db.execute("SELECT MAX(sort_order) as mx FROM payment_methods").fetchone()["mx"] or 0
        db.execute("INSERT INTO payment_methods (id, name, bank_group, is_active, sort_order) VALUES (?,?,?,1,?)",
                   (pid, body.name, body.bankGroup, mx+1))
    return {"ok": True, "id": pid}

@app.delete("/api/payment-methods/{pid}")
def delete_payment_method(pid: str, user=Depends(require_admin)):
    with get_db() as db:
        db.execute("UPDATE payment_methods SET is_active=0 WHERE id=?", (pid,))
    return {"ok": True}

# ══════════════════════════════════════════
# CASHFLOW REPORT
# ══════════════════════════════════════════
@app.get("/api/cashflow-report")
def cashflow_report(currency: str = "USD", user=Depends(require_admin)):
    """Monthly cashflow report with auto-calculated + manual override balances."""
    summary = compute(0, 0, 15000, True)
    rate = get_exchange_rate()
    
    # Load manual bank balances (now keyed by month+bank)
    with get_db() as db:
        bal_rows = db.execute("SELECT * FROM bank_balances ORDER BY month, bank").fetchall()
        pm_rows = db.execute("SELECT * FROM payment_methods WHERE is_active=1 ORDER BY sort_order").fetchall()
    
    # Group balances by month -> bank -> data
    manual_balances = {}
    for r in bal_rows:
        manual_balances.setdefault(r["month"], {})[r["bank"]] = dict(r)
    
    # Get unique bank groups from payment methods
    bank_groups = list(set(p["bank_group"] for p in pm_rows if p["bank_group"]))
    bank_groups.sort()
    
    now = datetime.now()
    current_month_str = f"{now.month:02d}-{now.year}"
    
    months = []
    prev_closing = 0
    prev_bank_closing = {bg: 0 for bg in bank_groups}
    
    for i, s in enumerate(summary):
        ml = ML[i]
        is_past = ml < current_month_str
        is_current = ml == current_month_str
        
        revenue = s["cashIn"]
        labor_total = sum(s["lab"].values())
        vendor_total = sum(s["vd"].values())
        total_expenses = labor_total + vendor_total
        net = revenue - total_expenses
        
        auto_opening = prev_closing
        auto_closing = auto_opening + net
        
        manual_month = manual_balances.get(ml, {})
        manual_total = manual_month.get("total", {})
        has_manual = manual_total.get("is_manual", 0)
        
        opening = manual_total.get("opening_balance") if (has_manual and manual_total.get("opening_balance") is not None) else auto_opening
        closing = manual_total.get("closing_balance") if (has_manual and manual_total.get("closing_balance") is not None) else (opening + net)
        
        # Per-bank balances
        banks_data = {}
        for bg in bank_groups:
            bank_manual = manual_month.get(bg, {})
            bk_has_manual = bank_manual.get("is_manual", 0)
            bk_auto_open = prev_bank_closing.get(bg, 0)
            bk_open = bank_manual.get("opening_balance") if (bk_has_manual and bank_manual.get("opening_balance") is not None) else bk_auto_open
            bk_close = bank_manual.get("closing_balance") if (bk_has_manual and bank_manual.get("closing_balance") is not None) else bk_open
            banks_data[bg] = {"opening": round(bk_open, 2), "closing": round(bk_close, 2), "isManual": bool(bk_has_manual)}
            prev_bank_closing[bg] = bk_close if bk_has_manual else bk_auto_open
        
        conv = rate if currency == "ILS" else 1.0
        
        month_data = {
            "month": ml,
            "monthLabel": ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][i],
            "isPast": is_past, "isCurrent": is_current,
            "isForecast": not is_past and not is_current,
            "openingBalance": round(opening * conv, 2),
            "revenue": round(revenue * conv, 2),
            "laborCost": round(labor_total * conv, 2),
            "vendorExpenses": round(vendor_total * conv, 2),
            "totalExpenses": round(total_expenses * conv, 2),
            "netCashflow": round(net * conv, 2),
            "closingBalance": round(closing * conv, 2),
            "isManual": bool(has_manual),
            "notes": manual_total.get("notes", ""),
            "openingBalanceUsd": round(opening, 2),
            "closingBalanceUsd": round(closing, 2),
            "banks": {bg: {"opening": round(banks_data[bg]["opening"] * conv, 2), "closing": round(banks_data[bg]["closing"] * conv, 2), "isManual": banks_data[bg]["isManual"]} for bg in bank_groups},
            # Payroll breakdown for cash flow
            "cfGross": round(s.get("cfGross", 0) * conv, 2),
            "cfSS": round(s.get("cfSS", 0) * conv, 2),
            "cfPension": round(s.get("cfPension", 0) * conv, 2),
            "cfStudyFund": round(s.get("cfStudyFund", 0) * conv, 2),
            "cfDisability": round(s.get("cfDisability", 0) * conv, 2),
            "cfTax": round(s.get("cfTax", 0) * conv, 2),
            "loanInterest": round(s.get("loanInterest", 0) * conv, 2),
            "loanPrincipal": round(s.get("loanPrincipal", 0) * conv, 2),
            "loanPayment": round(s.get("loanPayment", 0) * conv, 2),
            # Detailed breakdowns
            "labByDept": {k: round(v * conv, 2) for k, v in s["lab"].items() if v > 0},
            "vdByDept": {k: round(v * conv, 2) for k, v in s["vd"].items() if v > 0},
        }
        months.append(month_data)
        prev_closing = closing
    
    return {
        "months": months,
        "currency": currency,
        "exchangeRate": rate,
        "currentMonth": current_month_str,
        "bankGroups": bank_groups,
        "paymentMethods": [{"id":p["id"],"name":p["name"],"bankGroup":p["bank_group"]} for p in pm_rows]
    }

class BalanceUpdate(BaseModel):
    month: str
    bank: str = "total"
    openingBalance: Optional[float] = None
    closingBalance: Optional[float] = None
    notes: str = ""

@app.put("/api/cashflow-report/balance")
def update_balance(b: BalanceUpdate, user=Depends(require_admin)):
    with get_db() as db:
        existing = db.execute("SELECT * FROM bank_balances WHERE month=? AND bank=?", (b.month, b.bank)).fetchone()
        if existing:
            fields = ["is_manual=1"]
            vals = []
            if b.openingBalance is not None:
                fields.append("opening_balance=?"); vals.append(b.openingBalance)
            if b.closingBalance is not None:
                fields.append("closing_balance=?"); vals.append(b.closingBalance)
            if b.notes:
                fields.append("notes=?"); vals.append(b.notes)
            vals.extend([b.month, b.bank])
            db.execute(f"UPDATE bank_balances SET {','.join(fields)} WHERE month=? AND bank=?", vals)
        else:
            db.execute("INSERT INTO bank_balances (month, bank, opening_balance, closing_balance, is_manual, notes) VALUES (?,?,?,?,1,?)",
                       (b.month, b.bank, b.openingBalance, b.closingBalance, b.notes))
    return {"ok": True}

@app.post("/api/cashflow-report/upload")
async def upload_bank_statement(file: UploadFile, user=Depends(require_admin)):
    """Parse bank Excel file and extract monthly opening/closing balances."""
    import io
    try:
        import openpyxl
    except:
        return {"ok": False, "error": "openpyxl not installed"}
    
    data = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    
    # Try to find date and balance columns
    results = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        results.append([str(c) if c else "" for c in row])
    
    return {"ok": True, "rows": len(results), "preview": results[:10], "columns": results[0] if results else []}

# ══════════════════════════════════════════
# Serve
# ══════════════════════════════════════════
@app.get("/", response_class=HTMLResponse)
def serve():
    return (Path(__file__).parent / "index.html").read_text(encoding="utf-8")

@app.on_event("startup")
def startup():
    init_db()
    auto_migrate()
    if not is_seeded():
        print("  \U0001f331 Seeding..."); seed_db()
        print(f"  \u2705 Done")
    else:
        print(f"  \U0001f4be DB loaded")
    lock_past_months()
    now = datetime.now()
    print(f"  \U0001f4b1 Rate: $1 = \u20aa{get_exchange_rate():.2f} (locked for months before {now.month:02d}-{now.year})")
    print(f"\n  Users: admin/admin123 (management), eduardo/sales123, elad/sales123, tiki/sales123 (sales)\n")

if __name__ == "__main__":
    import uvicorn
    print("\n  \u26a1 FinStack v6 — Auth + CRM")
    print("  " + "\u2500" * 35)
    print("  http://localhost:8000\n")
    uvicorn.run(app, host="127.0.0.1", port=8000)
